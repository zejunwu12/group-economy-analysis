"""End-to-end regression tests for next-quarter template generation."""

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Color, PatternFill

from engine.config_loader import ConfigLoader
from engine.next_template import (
    NextTemplateConsistencyError,
    NextTemplateError,
    generate_next_template,
)
from engine.period import QuarterContext


class NextTemplateGenerationTests(unittest.TestCase):
    def _config(self, directory: str) -> dict:
        config_path = Path(__file__).resolve().parents[1] / "config.yaml"
        config = ConfigLoader(str(config_path)).load()
        config["runtime"]["_config_dir"] = directory
        config["runtime"]["output_dir"] = "output"
        return config

    def _source_workbook(self, path: Path, config: dict) -> None:
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = config["reports"]["report1"]["sheet_name"]
        for report_id in range(2, 9):
            workbook.create_sheet(
                config["reports"][f"report{report_id}"]["sheet_name"]
            )

        report1 = workbook[config["reports"]["report1"]["sheet_name"]]
        report1["A2"] = "文旅集团2026年第二季度资产总体情况表"
        report1["A3"] = "表头合并"
        report1.merge_cells("A3:B3")
        report1["A8"] = "数据区合并"
        report1.merge_cells("A8:A9")
        report1["C7"] = 100
        report1["D7"] = 200
        report1["E7"] = 150
        report1["F7"] = 120
        report1["G7"] = "=D7-E7"
        report1["I7"] = 90
        report1["I7"].comment = Comment("本期说明", "填报人")
        report1["C17"] = "=SUM(C7:C16)"
        report1["A20"] = "2026年第二季度比2026年第一季度的差异（环比）"
        report1["A21"] = "2026年第二季度同比2025年第二季度的差异（同比）"
        for cell in report1[16][:33]:
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

        report2 = workbook[config["reports"]["report2"]["sheet_name"]]
        report2["C4"] = 10
        report2["C4"].fill = PatternFill(
            fill_type="solid",
            fgColor=Color(theme=1),
        )
        report2["C14"] = "=SUM(C4:C13)"

        report3 = workbook[config["reports"]["report3"]["sheet_name"]]
        report3["A2"] = "文旅集团2026年第二季度资产运营管理情况"
        report3["C5"] = 10
        report3["D5"] = 8
        report3["F5"] = 20
        report3["G5"] = 18
        report3["I5"] = 30
        report3["J5"] = 28
        report3["L5"] = "=I5/J5-1"
        report3["N5"] = "本期原因"
        report3.merge_cells("D14:X14")
        report3["C15"] = "=SUM(C5:C14)"

        report4 = workbook[config["reports"]["report4"]["sheet_name"]]
        report4["C4"] = 10
        report4["H4"] = "=IFERROR(F4/C4,\"/\")"
        report4["C14"] = "=SUM(C4:C13)"

        report5 = workbook[config["reports"]["report5"]["sheet_name"]]
        report5["A2"] = "2026年第二季度项目改造情况（截止6月30日）"
        report5["A5"] = 1
        report5["B5"] = "拟改造项目"
        report5["F5"] = 1
        report5["G5"] = "正在改造项目"
        report5["A70"] = "合计"
        report5["D70"] = "=SUM(D5:D69)"
        report5["F70"] = "合计"
        report5["I70"] = "=SUM(I5:I69)"
        report5.merge_cells("A70:C70")
        report5.merge_cells("F70:H70")

        report6 = workbook[config["reports"]["report6"]["sheet_name"]]
        report6["C6"] = 10
        report6["C16"] = "=SUM(C6:C15)"

        report7 = workbook[config["reports"]["report7"]["sheet_name"]]
        report7["A2"] = "园区2026年第二季度招商运营情况表"
        report7["A4"] = "测试园区"
        report7["B4"] = "资产主方"
        report7["C4"] = "运营公司"
        report7["D4"] = 100
        report7["N4"] = "=J4/I4"
        report7["D8"] = "=SUM(D4:D7)"

        report8 = workbook[config["reports"]["report8"]["sheet_name"]]
        report8["A3"] = "日期：2026年6月30日"
        report8["A5"] = 1
        report8["B5"] = "非自有资产"
        report8["K12"] = "跨行说明"
        report8.merge_cells("K12:K14")
        report8["K15"] = "跨越缩行边界"
        report8.merge_cells("K15:K21")
        report8["K22"] = "随删除行移除"
        report8.merge_cells("K22:K23")
        report8["A158"] = 154
        report8["B158"] = "扩展行资产"

        for report_id in (1, 2, 3, 4, 6):
            report_config = config["reports"][f"report{report_id}"]
            worksheet = workbook[report_config["sheet_name"]]
            for row, unit_name in report_config["row_mapping"].items():
                worksheet[f"{report_config['b_col']}{row}"] = unit_name
            worksheet[
                f"{report_config.get('a_col', 'A')}{report_config['total_row']}"
            ] = "合计"

        report7_config = config["reports"]["report7"]
        for sub_table in report7_config["sub_tables"]:
            for row, park_name in sub_table["row_mapping"].items():
                report7[f"A{row}"] = park_name
            if isinstance(sub_table.get("total_row"), int):
                report7[f"A{sub_table['total_row']}"] = "合计"

        workbook.save(path)
        workbook.close()

    def test_generates_next_template_with_roll_clear_and_structure_rules(self):
        with tempfile.TemporaryDirectory() as directory:
            config = self._config(directory)
            source_path = Path(directory) / "final.xlsx"
            self._source_workbook(source_path, config)

            result = generate_next_template(
                source_path,
                QuarterContext.parse("2026Q2"),
                config,
            )

            output_path = Path(result.output_path)
            self.assertEqual(output_path.parent, Path(directory) / "output")
            self.assertIn("2026年第三季度", output_path.name)
            self.assertEqual(result.source_quarter, "2026Q2")
            self.assertEqual(result.target_quarter, "2026Q3")

            source = load_workbook(source_path, data_only=False)
            target = load_workbook(output_path, data_only=False)
            try:
                report1 = target[config["reports"]["report1"]["sheet_name"]]
                self.assertIn("2026年第三季度", report1["A2"].value)
                self.assertEqual(report1["C7"].value, 100)
                self.assertIsNone(report1["D7"].value)
                self.assertEqual(report1["E7"].value, 200)
                self.assertEqual(report1["G7"].value, "=D7-E7")
                self.assertIsNone(report1["I7"].value)
                self.assertTrue(
                    all(cell.fill.fill_type is None for cell in report1[16][:33])
                )
                self.assertIn("A3:B3", map(str, report1.merged_cells.ranges))
                self.assertIn("A8:A9", map(str, report1.merged_cells.ranges))
                self.assertEqual(
                    report1["A20"].value,
                    "2026年第三季度比2026年第二季度的差异（环比）",
                )
                self.assertEqual(
                    report1["A21"].value,
                    "2026年第三季度同比2025年第三季度的差异（同比）",
                )

                report2 = target[config["reports"]["report2"]["sheet_name"]]
                self.assertEqual(report2["C4"].fill.fill_type, "solid")
                self.assertEqual(report2["C4"].fill.fgColor.type, "theme")

                report3 = target[config["reports"]["report3"]["sheet_name"]]
                self.assertEqual(report3["D5"].value, 10)
                self.assertEqual(report3["G5"].value, 20)
                self.assertEqual(report3["J5"].value, 30)
                self.assertIsNone(report3["C5"].value)
                self.assertEqual(report3["L5"].value, "=I5/J5-1")
                self.assertIsNone(report3["N5"].value)
                self.assertIn("D14:X14", map(str, report3.merged_cells.ranges))

                report5 = target[config["reports"]["report5"]["sheet_name"]]
                self.assertEqual(report5.max_row, 20)
                self.assertEqual(report5["A5"].value, 1)
                self.assertIsNone(report5["B5"].value)
                self.assertEqual(report5["A19"].value, 15)
                self.assertEqual(report5["A20"].value, "合计")
                self.assertEqual(report5["D20"].value, "=SUM(D5:D19)")
                self.assertEqual(report5["I20"].value, "=SUM(I5:I19)")
                self.assertIn("A20:C20", map(str, report5.merged_cells.ranges))
                self.assertIn("F20:H20", map(str, report5.merged_cells.ranges))

                report7 = target[config["reports"]["report7"]["sheet_name"]]
                self.assertEqual(
                    report7["A4"].value,
                    config["reports"]["report7"]["sub_tables"][0]["row_mapping"][4],
                )
                self.assertIsNone(report7["D4"].value)
                self.assertEqual(report7["N4"].value, "=J4/I4")

                report8 = target[config["reports"]["report8"]["sheet_name"]]
                self.assertEqual(report8.max_row, 20)
                self.assertEqual(report8["A20"].value, 16)
                self.assertIsNone(report8["B20"].value)
                self.assertIn("K12:K14", map(str, report8.merged_cells.ranges))
                self.assertIn("K15:K20", map(str, report8.merged_cells.ranges))
                self.assertNotIn("K22:K23", map(str, report8.merged_cells.ranges))
                self.assertIsInstance(report8["K13"], MergedCell)
                self.assertIsInstance(report8["K19"], MergedCell)

                for report_id in (1, 2, 3, 4, 6, 7):
                    sheet_name = config["reports"][f"report{report_id}"]["sheet_name"]
                    self.assertEqual(
                        tuple(sorted(map(str, source[sheet_name].merged_cells.ranges))),
                        tuple(sorted(map(str, target[sheet_name].merged_cells.ranges))),
                    )

                self.assertTrue(
                    all(
                        cell.comment is None
                        for worksheet in target.worksheets
                        for row in worksheet.iter_rows()
                        for cell in row
                    )
                )
            finally:
                source.close()
                target.close()

    def test_rejects_source_config_drift_before_writing(self):
        with tempfile.TemporaryDirectory() as directory:
            config = self._config(directory)
            source_path = Path(directory) / "final.xlsx"
            self._source_workbook(source_path, config)

            workbook = load_workbook(source_path, data_only=False)
            report1 = workbook[config["reports"]["report1"]["sheet_name"]]
            report1.insert_rows(12)
            report1["A12"] = "酒管集团"
            report1["B12"] = "泉旅酒管"
            workbook.save(source_path)
            workbook.close()

            with self.assertLogs("engine.next_template", level="WARNING") as captured:
                with self.assertRaises(NextTemplateConsistencyError) as caught:
                    generate_next_template(
                        source_path,
                        QuarterContext.parse("2026Q2"),
                        config,
                    )

            warning_text = "\n".join(captured.output)
            self.assertIn("新增单位", warning_text)
            self.assertIn("泉旅酒管", warning_text)
            self.assertIn("单位行位移", warning_text)
            self.assertIn("合计行变化", warning_text)
            self.assertIn("数据区结束行变化", warning_text)
            self.assertGreaterEqual(len(caught.exception.issues), 5)

            output_dir = Path(directory) / "output"
            self.assertFalse(output_dir.exists())

            unchanged = load_workbook(source_path, data_only=False)
            try:
                self.assertIsNotNone(
                    unchanged[config["reports"]["report1"]["sheet_name"]]["I7"].comment
                )
            finally:
                unchanged.close()

    def test_rejects_source_when_quarter_marker_does_not_match(self):
        with tempfile.TemporaryDirectory() as directory:
            config = self._config(directory)
            source_path = Path(directory) / "final.xlsx"
            self._source_workbook(source_path, config)

            with self.assertRaisesRegex(NextTemplateError, "请核对 --quarter"):
                generate_next_template(
                    source_path,
                    QuarterContext.parse("2025Q4"),
                    config,
                )


if __name__ == "__main__":
    unittest.main()

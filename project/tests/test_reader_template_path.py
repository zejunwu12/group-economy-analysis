"""Template path default and runtime override tests."""

from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from engine.reader import load_template


class TemplatePathTests(unittest.TestCase):
    def _config(self, config_dir: Path, template_path: str) -> dict:
        return {
            "runtime": {
                "_config_dir": str(config_dir),
                "template_path": template_path,
            }
        }

    def test_uses_config_template_path_when_override_is_missing(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config_dir = root / "config"
            config_dir.mkdir()
            default_path = config_dir / "default.xlsx"
            default_path.touch()

            with patch("engine.reader.openpyxl.load_workbook", return_value=Workbook()) as load:
                workbook = load_template(
                    self._config(config_dir, "default.xlsx")
                )

            self.assertIsNotNone(workbook)
            self.assertEqual(load.call_args.args[0], str(default_path.resolve()))

    def test_runtime_template_override_takes_precedence(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config_dir = root / "config"
            config_dir.mkdir()
            default_path = config_dir / "default.xlsx"
            override_path = root / "final.xlsx"
            default_path.touch()
            override_path.touch()

            with patch("engine.reader.openpyxl.load_workbook", return_value=Workbook()) as load:
                workbook = load_template(
                    self._config(config_dir, "default.xlsx"),
                    template_path=override_path,
                )

            self.assertIsNotNone(workbook)
            self.assertEqual(load.call_args.args[0], str(override_path.resolve()))

    def test_rejects_blank_runtime_override(self):
        with tempfile.TemporaryDirectory() as directory:
            config_dir = Path(directory)
            with self.assertRaisesRegex(ValueError, "模板路径不能为空"):
                load_template(
                    self._config(config_dir, "default.xlsx"),
                    template_path="   ",
                )


if __name__ == "__main__":
    unittest.main()

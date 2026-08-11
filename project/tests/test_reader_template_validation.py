"""写入前模板与配置结构校验。"""

import unittest

from openpyxl import Workbook

from engine.reader import TemplateMismatchError, validate_template


class DynamicReportTemplateValidationTests(unittest.TestCase):
    def _workbook(self) -> Workbook:
        workbook = Workbook()
        report5 = workbook.active
        report5.title = "报表5"
        report5["J25"] = None
        report5["D25"] = "=SUM(D5:D24)"
        report5["I25"] = "=SUM(I5:I24)"
        report5.merge_cells("A25:C25")
        report5.merge_cells("F25:H25")

        report8 = workbook.create_sheet("报表8")
        report8["L24"] = None
        report8.merge_cells("A1:L1")
        report8.merge_cells("K14:K20")
        return workbook

    def _config(self) -> dict:
        return {
            "reports": {
                "report5": {
                    "sheet_name": "报表5",
                    "left": {
                        "data_start_row": 5,
                        "data_end_row": 24,
                        "total_row": 25,
                        "cols": ["A", "B", "C", "D", "E"],
                    },
                    "right": {
                        "data_start_row": 5,
                        "data_end_row": 24,
                        "total_row": 25,
                        "cols": ["F", "G", "H", "I", "J"],
                    },
                },
                "report8": {
                    "sheet_name": "报表8",
                    "data_start_row": 5,
                    "data_end_row": 24,
                    "cols": [
                        "A", "B", "C", "D", "E", "F",
                        "G", "H", "I", "J", "K", "L",
                    ],
                },
            }
        }

    def test_accepts_matching_report5_and_report8_structures(self):
        workbook = self._workbook()
        try:
            passed = validate_template(workbook, self._config(), report_ids=[5, 8])
        finally:
            workbook.close()

        self.assertEqual(passed, ["报表5", "报表8"])

    def test_report5_reports_old_row_config_and_merged_cell_conflicts(self):
        workbook = self._workbook()
        config = self._config()
        for side in ("left", "right"):
            config["reports"]["report5"][side]["data_end_row"] = 63
            config["reports"]["report5"][side]["total_row"] = 64
        merges_before = set(map(str, workbook["报表5"].merged_cells.ranges))

        try:
            with self.assertRaises(TemplateMismatchError) as raised:
                validate_template(workbook, config, report_ids=[5])
            merges_after = set(map(str, workbook["报表5"].merged_cells.ranges))
        finally:
            workbook.close()

        message = str(raised.exception)
        self.assertIn("模板与配置前置校验失败", message)
        self.assertIn("reports.report5.left.data_end_row=63", message)
        self.assertIn("reports.report5.left.total_row=64", message)
        self.assertIn("合并区域 A25:C25", message)
        self.assertIn("合并区域 F25:H25", message)
        self.assertIn("处理尚未开始，模板未被修改", message)
        self.assertEqual(merges_after, merges_before)

    def test_report8_reports_data_end_row_mismatch_before_writing(self):
        workbook = self._workbook()
        config = self._config()
        config["reports"]["report8"]["data_end_row"] = 155

        try:
            with self.assertRaises(TemplateMismatchError) as raised:
                validate_template(workbook, config, report_ids=[8])
        finally:
            workbook.close()

        message = str(raised.exception)
        self.assertIn("reports.report8.data_end_row=155", message)
        self.assertIn("实际最大行 24", message)

    def test_report8_reports_merge_crossing_header_boundary(self):
        workbook = self._workbook()
        workbook["报表8"].merge_cells("J4:J5")

        try:
            with self.assertRaises(TemplateMismatchError) as raised:
                validate_template(workbook, self._config(), report_ids=[8])
        finally:
            workbook.close()

        message = str(raised.exception)
        self.assertIn("合并区域 J4:J5", message)
        self.assertIn("跨越表头和数据区边界", message)

    def test_collects_report5_and_report8_problems_in_one_error(self):
        workbook = self._workbook()
        config = self._config()
        for side in ("left", "right"):
            config["reports"]["report5"][side]["data_end_row"] = 63
            config["reports"]["report5"][side]["total_row"] = 64
        config["reports"]["report8"]["data_end_row"] = 155

        try:
            with self.assertRaises(TemplateMismatchError) as raised:
                validate_template(workbook, config, report_ids=[5, 8])
        finally:
            workbook.close()

        message = str(raised.exception)
        self.assertIn("报表5 左栏", message)
        self.assertIn("报表8: reports.report8.data_end_row=155", message)
        self.assertRegex(message, r"共发现 \d+ 项")


if __name__ == "__main__":
    unittest.main()

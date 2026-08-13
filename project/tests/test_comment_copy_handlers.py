"""覆盖明细和园区报表的权属批注复制。"""

from io import BytesIO
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side

from engine.comments import CommentCopyStats, clear_template_comments
from engine.report5_handler import process_report5
from engine.report7_handler import process_report7
from engine.report8_handler import process_report8


class CommentCopyHandlerTests(unittest.TestCase):
    def test_template_comments_are_cleared_and_counted(self):
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "报表1"
        first_sheet["C4"].comment = Comment("模板遗留", "Lenovo")
        second_sheet = workbook.create_sheet("报表7")
        second_sheet["S6"].comment = Comment("旧口径", "Administrator")
        stats = CommentCopyStats()

        with self.assertLogs("engine.comments", level="DEBUG") as captured:
            cleared = clear_template_comments(workbook, stats=stats)

        self.assertEqual(cleared, 2)
        self.assertEqual(stats.template_cleared, 2)
        self.assertIsNone(first_sheet["C4"].comment)
        self.assertIsNone(second_sheet["S6"].comment)
        log_text = "\n".join(captured.output)
        self.assertIn("模板批注已清除", log_text)
        self.assertIn("报表7!S6", log_text)
        self.assertIn("共清除 2 条", log_text)

    def test_report5_copies_detail_comment(self):
        template_book = Workbook()
        target = template_book.active
        target.title = "报表5"
        target["E6"] = "=SUM(E5:E5)"
        target["J6"] = "=SUM(J5:J5)"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表5"
        source["A5"] = 1
        source["B5"] = "改造项目"
        source["B5"].comment = Comment("改造范围以现场确认为准", "填报人")

        stats = CommentCopyStats()
        process_report5(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表5",
                "left": {
                    "data_start_row": 5, "data_end_row": 5,
                    "total_row": 6, "cols": ["A", "B", "C", "D", "E"],
                },
                "right": {
                    "data_start_row": 5, "data_end_row": 5,
                    "total_row": 6, "cols": ["F", "G", "H", "I", "J"],
                },
            },
            stats,
        )

        self.assertEqual(target["B5"].comment.text, "改造范围以现场确认为准")
        self.assertEqual(target["B5"].comment.author, "填报人")
        self.assertEqual(stats.copied, 1)

    def test_report5_excludes_failed_owner(self):
        template_book = Workbook()
        target = template_book.active
        target.title = "报表5"
        target["E6"] = "=SUM(E5:E5)"
        target["J6"] = "=SUM(J5:J5)"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表5"
        source["A5"] = 1
        source["B5"] = "不应写入"

        result = process_report5(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表5",
                "left": {
                    "data_start_row": 5, "data_end_row": 5,
                    "total_row": 6, "cols": ["A", "B", "C", "D", "E"],
                },
                "right": {
                    "data_start_row": 5, "data_end_row": 5,
                    "total_row": 6, "cols": ["F", "G", "H", "I", "J"],
                },
            },
            excluded_owners={"权属A"},
        )

        self.assertEqual(result["left_count"], 0)
        self.assertIsNone(target["B5"].value)

    def test_report7_copies_park_comment(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表7"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表7"
        source["A4"] = "测试园区"
        source["D4"] = 10
        source["D4"].comment = Comment("出租面积含临时合同", "管理员")

        stats = CommentCopyStats()
        process_report7(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表7", "data_start_col": "A", "data_end_col": "D",
                "sub_tables": [{"row_mapping": {4: "测试园区"}}],
            },
            stats,
        )

        self.assertEqual(target["D4"].comment.text, "出租面积含临时合同")
        self.assertEqual(stats.copied, 1)

    def test_report7_excludes_failed_owner_from_source_search(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表7"

        failed_book = Workbook()
        failed = failed_book.active
        failed.title = "报表7"
        failed["A4"] = "测试园区"
        failed["D4"] = 99

        valid_book = Workbook()
        valid = valid_book.active
        valid.title = "报表7"
        valid["A4"] = "测试园区"
        valid["D4"] = 10

        result = process_report7(
            target,
            {
                "异常权属": {"workbook": failed_book},
                "正常权属": {"workbook": valid_book},
            },
            {
                "sheet_name": "报表7", "data_start_col": "A", "data_end_col": "D",
                "sub_tables": [{"row_mapping": {4: "测试园区"}}],
            },
            excluded_owners={"异常权属"},
        )

        self.assertEqual(result["测试园区"]["owner"], "正常权属")
        self.assertEqual(target["D4"].value, 10)

    def test_report7_preserves_only_configured_formula_columns_and_total_rows(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表7"
        target["A4"] = "园区甲"
        target["A23"] = "园区乙"
        for row in (4, 23):
            target[f"F{row}"] = f"=D{row}-E{row}-I{row}"
            target[f"J{row}"] = f"=K{row}+1"
            target[f"N{row}"] = f"=J{row}/I{row}"
            target[f"O{row}"] = f"=K{row}/I{row}"
            target[f"R{row}"] = f"=P{row}/Q{row}-1"
            target[f"T{row}"] = f"=P{row}/S{row}-1"
        target["D8"] = "=SUM(D4:D7)"
        target["D26"] = "=SUM(D23:D25)"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表7"
        for row, park_name, offset in ((4, "园区甲", 0), (23, "园区乙", 100)):
            source[f"A{row}"] = park_name
            source[f"D{row}"] = 10 + offset
            source[f"F{row}"] = 20 + offset
            source[f"J{row}"] = 30 + offset
            source[f"N{row}"] = 0.4
            source[f"O{row}"] = 0.3
            source[f"R{row}"] = 0.2
            source[f"T{row}"] = 0.1

        results = process_report7(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表7",
                "data_start_col": "A",
                "data_end_col": "V",
                "formula_columns": ["N", "O", "R", "T"],
                "text_columns": ["B", "C", "U", "V"],
                "sub_tables": [
                    {
                        "total_row": 8,
                        "row_mapping": {4: "园区甲"},
                    },
                    {
                        "total_row": 26,
                        "row_mapping": {23: "园区乙"},
                    },
                ],
            },
        )

        for row, offset in ((4, 0), (23, 100)):
            self.assertEqual(target[f"F{row}"].value, 20 + offset)
            self.assertEqual(target[f"J{row}"].value, 30 + offset)
            self.assertEqual(target[f"N{row}"].value, f"=J{row}/I{row}")
            self.assertEqual(target[f"O{row}"].value, f"=K{row}/I{row}")
            self.assertEqual(target[f"R{row}"].value, f"=P{row}/Q{row}-1")
            self.assertEqual(target[f"T{row}"].value, f"=P{row}/S{row}-1")
        self.assertEqual(target["D8"].value, "=SUM(D4:D7)")
        self.assertEqual(target["D26"].value, "=SUM(D23:D25)")
        self.assertEqual(set(results), {"园区甲", "园区乙"})

    def test_report7_zero_fills_non_numeric_cells_and_preserves_text(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表7"
        target["A4"] = "测试园区"
        for column in ("B", "C", "D", "E", "F", "G", "U", "V"):
            target[f"{column}4"] = "模板旧值"
        target["N4"] = "=J4/I4"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表7"
        source["A4"] = "测试园区"
        source["B4"] = "资产主方"
        source["C4"] = None
        source["D4"] = 10
        source["E4"] = None
        source["F4"] = "/"
        source["G4"] = "待确认"
        source["N4"] = 0.5
        source["U4"] = "差异原因"
        source["V4"] = None

        process_report7(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表7",
                "data_start_col": "A",
                "data_end_col": "V",
                "formula_columns": ["N", "O", "R", "T"],
                "text_columns": ["B", "C", "U", "V"],
                "sub_tables": [{"row_mapping": {4: "测试园区"}}],
            },
        )

        self.assertEqual(target["B4"].value, "资产主方")
        self.assertIsNone(target["C4"].value)
        self.assertEqual(target["D4"].value, 10)
        self.assertEqual(target["E4"].value, 0)
        self.assertEqual(target["F4"].value, 0)
        self.assertEqual(target["G4"].value, 0)
        self.assertEqual(target["N4"].value, "=J4/I4")
        self.assertEqual(target["U4"].value, "差异原因")
        self.assertIsNone(target["V4"].value)

    def test_report8_copies_detail_comment_but_not_replaced_sequence(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表8"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表8"
        source["A5"] = 99
        source["A5"].comment = Comment("来源序号", "填报人")
        source["B5"] = "非自有资产"
        source["B5"].comment = Comment("租赁期限待复核", "填报人")

        stats = CommentCopyStats()
        process_report8(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表8", "data_start_row": 5,
                "data_end_row": 5, "cols": ["A", "B", "C"],
            },
            stats,
        )

        self.assertEqual(target["A5"].value, 1)
        self.assertIsNone(target["A5"].comment)
        self.assertEqual(target["B5"].comment.text, "租赁期限待复核")
        self.assertEqual(stats.copied, 1)

    def test_report8_registers_cross_workbook_styles_before_recreating_merge(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "报表8"

        source_book = Workbook()
        source = source_book.active
        source.title = "报表8"
        for index in range(40):
            source.cell(row=1, column=13 + index).border = Border(
                left=Side(style="thin", color=f"FF{index:06X}")
            )
        for row in range(40, 45):
            source.cell(row=row, column=1).value = row
            source.cell(row=row, column=2).value = f"非自有资产{row}"
        source["K40"] = "合并说明"
        source["K44"].border = Border(
            bottom=Side(style="thin", color="FFFF0000")
        )
        source.merge_cells("K40:K44")

        process_report8(
            target,
            {"权属A": {"workbook": source_book}},
            {
                "sheet_name": "报表8",
                "data_start_row": 5,
                "data_end_row": 5,
                "cols": [
                    "A", "B", "C", "D", "E", "F",
                    "G", "H", "I", "J", "K", "L",
                ],
            },
        )

        self.assertIn("K5:K9", map(str, target.merged_cells.ranges))
        self.assertEqual(target["B5"].value, "非自有资产40")

        output = BytesIO()
        target_book.save(output)
        output.seek(0)
        reloaded = load_workbook(output)
        try:
            reloaded_target = reloaded["报表8"]
            self.assertIn("K5:K9", map(str, reloaded_target.merged_cells.ranges))
        finally:
            reloaded.close()

    def test_comment_stats_rollback_removes_copied_details(self):
        target_book = Workbook()
        target = target_book.active
        target.title = "汇总表"
        source_book = Workbook()
        source = source_book.active
        source.title = "权属表"
        source["C4"].comment = Comment("待回滚批注", "填报人")
        stats = CommentCopyStats()
        checkpoint = stats.checkpoint()

        from engine.comments import copy_source_comment

        copy_source_comment(
            source["C4"],
            target["C4"],
            report_id=1,
            owner_key="权属A",
            source_ws=source,
            target_ws=target,
            stats=stats,
        )
        removed = stats.rollback(checkpoint)

        self.assertEqual(removed, 1)
        self.assertEqual(stats.copied, 0)
        self.assertEqual(stats.details, [])


if __name__ == "__main__":
    unittest.main()

"""Ownership header structure validation tests."""

import unittest

from openpyxl import Workbook

from engine.header_validator import detect_ownership_header_mismatches
from engine.report8_handler import process_report8


class HeaderValidatorTests(unittest.TestCase):
    def _workbook(self, *, columns: int = 3, merge: str | None = None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "报表8"
        for column in range(1, columns + 1):
            worksheet.cell(row=4, column=column).value = f"字段{column}"
        if merge:
            worksheet.merge_cells(merge)
        return workbook

    def _config(self):
        return {
            "reports": {
                "report8": {
                    "sheet_name": "报表8",
                    "header_start_row": 4,
                    "data_start_row": 5,
                    "data_end_row": 5,
                    "cols": ["A", "B", "C"],
                }
            }
        }

    def test_matching_header_passes(self):
        template = self._workbook(columns=3)
        source = self._workbook(columns=3)

        mismatches = detect_ownership_header_mismatches(
            template,
            {"权属A": {"workbook": source}},
            self._config(),
            report_ids=[8],
        )

        self.assertEqual(mismatches, {})

    def test_extra_source_column_is_reported(self):
        template = self._workbook(columns=3)
        source = self._workbook(columns=4)

        mismatches = detect_ownership_header_mismatches(
            template,
            {"权属A": {"workbook": source}},
            self._config(),
            report_ids=[8],
        )

        reason = mismatches[("权属A", 8)]
        self.assertIn("模板 A:C", reason)
        self.assertIn("权属表 A:D", reason)

    def test_merged_header_difference_is_reported(self):
        template = self._workbook(columns=3, merge="A4:B4")
        source = self._workbook(columns=3)

        mismatches = detect_ownership_header_mismatches(
            template,
            {"权属A": {"workbook": source}},
            self._config(),
            report_ids=[8],
        )

        self.assertIn("表头合并结构不一致", mismatches[("权属A", 8)])

    def test_missing_sheet_is_reported(self):
        template = self._workbook(columns=3)
        source = Workbook()
        source.active.title = "其他报表"

        mismatches = detect_ownership_header_mismatches(
            template,
            {"权属A": {"workbook": source}},
            self._config(),
            report_ids=[8],
        )

        self.assertIn("缺少工作表", mismatches[("权属A", 8)])

    def test_excluded_owner_is_not_written_by_report8(self):
        target = self._workbook(columns=3)
        source = self._workbook(columns=4)
        source["报表8"]["A5"] = 1
        source["报表8"]["B5"] = "不应写入"

        result = process_report8(
            target["报表8"],
            {"权属A": {"workbook": source}},
            self._config()["reports"]["report8"],
            excluded_owners={"权属A"},
        )

        self.assertEqual(result["record_count"], 0)
        self.assertIsNone(target["报表8"]["B5"].value)


if __name__ == "__main__":
    unittest.main()

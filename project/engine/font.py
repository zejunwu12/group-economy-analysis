"""Apply uniform font settings to data areas of the summary workbook."""

import logging

from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_FIXED_REPORT_IDS = frozenset({1, 2, 3, 4, 6})


def apply_uniform_font(workbook, config: dict) -> int:
    """按报表配置对数据区域应用统一字体。

    仅作用于各报表声明的数据区（含合计行），不影响标题和表头。
    支持按报表编号覆盖全局字体设置。

    Args:
        workbook: 已填充数据的汇总 Workbook
        config: 完整配置字典

    Returns:
        修改的单元格总数
    """
    font_config = config.get("font")
    if not isinstance(font_config, dict):
        logger.info("未配置 font 节，跳过统一字体")
        return 0
    if not font_config.get("enabled", False):
        logger.info("统一字体功能未启用（font.enabled=false）")
        return 0

    global_name = font_config.get("name", "微软雅黑")
    global_size = font_config.get("size", 10)
    global_bold = font_config.get("bold", False)
    report_overrides = font_config.get("report_overrides", {})
    if not isinstance(report_overrides, dict):
        report_overrides = {}

    reports_config = config.get("reports", {})
    modified_count = 0

    for report_key, report_config in reports_config.items():
        report_id = int(report_key.replace("report", ""))
        override = report_overrides.get(report_id, {})
        if not isinstance(override, dict):
            override = {}

        font_name = override.get("name", global_name)
        font_size = override.get("size", global_size)
        font_bold = override.get("bold", global_bold)
        target_font = Font(name=font_name, size=font_size, bold=font_bold)

        sheet_name = report_config.get("sheet_name", "")
        if not sheet_name or sheet_name not in workbook.sheetnames:
            logger.warning(
                f"报表{report_id}: 工作表 '{sheet_name}' 不存在，跳过字体统一"
            )
            continue

        worksheet = workbook[sheet_name]

        try:
            report_modified = _apply_to_report(
                worksheet,
                report_config,
                report_id,
                target_font,
            )
            modified_count += report_modified
            logger.debug(
                f"报表{report_id} '{sheet_name}': "
                f"统一字体 {font_name} {font_size}pt，"
                f"修改 {report_modified} 个单元格"
            )
        except Exception as exc:
            logger.warning(
                f"报表{report_id} 字体统一失败: {exc}，已跳过"
            )

    logger.info(
        f"统一字体完成: 共修改 {modified_count} 个单元格"
    )
    return modified_count


def _apply_to_report(
    worksheet: Worksheet,
    report_config: dict,
    report_id: int,
    font: Font,
) -> int:
    """根据报表类型确定数据区域并应用字体。"""
    if report_id in _FIXED_REPORT_IDS:
        return _apply_fixed_report(worksheet, report_config, font)
    if report_id == 5:
        return _apply_report5(worksheet, report_config, font)
    if report_id == 7:
        return _apply_report7(worksheet, report_config, font)
    if report_id == 8:
        return _apply_report8(worksheet, report_config, font)
    logger.warning(f"报表{report_id}: 不支持的类型，跳过字体统一")
    return 0


# ---------------------------------------------------------------------------
# Fixed reports (1, 2, 3, 4, 6)
# ---------------------------------------------------------------------------

def _apply_fixed_report(
    worksheet: Worksheet,
    report_config: dict,
    font: Font,
) -> int:
    """对固定行数报表的数据区和合计行应用字体。"""
    start_row = report_config["data_start_row"]
    end_row = report_config.get("total_row", report_config["data_end_row"])
    start_col = report_config["data_start_col"]
    end_col = report_config["data_end_col"]
    return _apply_font_range(worksheet, font, start_row, end_row, start_col, end_col)


# ---------------------------------------------------------------------------
# Report 5: double-column, dynamic row count
# ---------------------------------------------------------------------------

def _apply_report5(
    worksheet: Worksheet,
    report_config: dict,
    font: Font,
) -> int:
    """对报表5左右两栏数据区和合计行应用字体。

    合计行位置可能因动态扩展而变化，通过扫描 A 列定位。
    """
    left_config = report_config["left"]
    right_config = report_config["right"]
    data_start_row = left_config["data_start_row"]
    first_col = left_config["cols"][0]

    total_row = _find_total_row(worksheet, data_start_row, first_col)
    if total_row is None:
        logger.warning(
            "报表5: 未找到合计行，字体仅应用于数据区"
        )
        total_row = data_start_row  # fallback: no rows
        data_end = data_start_row - 1
    else:
        data_end = total_row - 1

    modified = 0
    # 左栏数据区
    modified += _apply_font_range(
        worksheet, font,
        data_start_row, data_end,
        left_config["cols"][0], left_config["cols"][-1],
    )
    # 右栏数据区
    modified += _apply_font_range(
        worksheet, font,
        data_start_row, data_end,
        right_config["cols"][0], right_config["cols"][-1],
    )
    # 合计行（左右两栏全宽）
    modified += _apply_font_range(
        worksheet, font,
        total_row, total_row,
        left_config["cols"][0], right_config["cols"][-1],
    )
    return modified


# ---------------------------------------------------------------------------
# Report 7: park sub-tables
# ---------------------------------------------------------------------------

def _apply_report7(
    worksheet: Worksheet,
    report_config: dict,
    font: Font,
) -> int:
    """对报表7各子表数据区和合计行应用字体。"""
    modified = 0
    start_col = report_config["data_start_col"]
    end_col = report_config["data_end_col"]

    for sub_table in report_config.get("sub_tables", []):
        sub_start = sub_table["data_start_row"]
        total_row = sub_table.get("total_row")
        if total_row is not None:
            # 数据区 + 合计行
            modified += _apply_font_range(
                worksheet, font,
                sub_start, total_row,
                start_col, end_col,
            )
        else:
            # 无合计行，仅数据区
            sub_end = sub_table["data_end_row"]
            modified += _apply_font_range(
                worksheet, font,
                sub_start, sub_end,
                start_col, end_col,
            )
    return modified


# ---------------------------------------------------------------------------
# Report 8: detail, dynamic row count
# ---------------------------------------------------------------------------

def _apply_report8(
    worksheet: Worksheet,
    report_config: dict,
    font: Font,
) -> int:
    """对报表8数据区应用字体。

    数据区可能因动态扩展而变化，通过扫描 A 列定位最后一行。
    """
    data_start_row = report_config["data_start_row"]
    first_col = report_config["cols"][0]
    last_col = report_config["cols"][-1]

    data_end_row = _find_last_data_row(worksheet, data_start_row, first_col)
    if data_end_row is None:
        logger.warning("报表8: 未找到数据行，跳过字体统一")
        return 0

    return _apply_font_range(
        worksheet, font,
        data_start_row, data_end_row,
        first_col, last_col,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_font_range(
    worksheet: Worksheet,
    font: Font,
    start_row: int,
    end_row: int,
    start_col_letter: str,
    end_col_letter: str,
) -> int:
    """对指定矩形区域内的单元格设置字体。

    Args:
        worksheet: 目标工作表
        font: 目标字体
        start_row / end_row: 行范围（含）
        start_col_letter / end_col_letter: 列范围（含），如 "A" / "K"

    Returns:
        修改的单元格数
    """
    if start_row > end_row:
        return 0

    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    count = 0

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = font
            count += 1

    return count


def _find_total_row(
    worksheet: Worksheet,
    start_row: int,
    col_letter: str,
    max_scan: int = 200,
) -> int | None:
    """扫描指定列，定位第一个包含"合计"的单元格所在行。

    对标 writer._is_total_row_label 的归一化逻辑。
    """
    col_idx = column_index_from_string(col_letter)
    for row in range(start_row, start_row + max_scan):
        value = worksheet.cell(row=row, column=col_idx).value
        if value is None:
            continue
        normalized = "".join(str(value).split())
        if _is_total_label(normalized):
            return row
    return None


def _find_last_data_row(
    worksheet: Worksheet,
    start_row: int,
    col_letter: str,
    max_scan: int = 500,
) -> int | None:
    """扫描指定列，定位最后一个非空数据行。

    从 start_row 开始扫描，遇到连续空行或非数字序号时停止。
    """
    col_idx = column_index_from_string(col_letter)
    last_data_row = None

    for row in range(start_row, start_row + max_scan):
        value = worksheet.cell(row=row, column=col_idx).value
        if value is None:
            break
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            last_data_row = row
        else:
            # 非数字值（如文本），视作数据区结束
            break

    return last_data_row


def _is_total_label(value: str) -> bool:
    """识别"合计""汇总""总计"等标记。"""
    total_labels = frozenset({"合计", "汇总", "总计"})
    if value in total_labels:
        return True
    return any(value.endswith(label) for label in total_labels)
"""Template and data reader"""

import logging
import os
from pathlib import Path
import openpyxl
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


class TemplateMismatchError(Exception):
    """模板与配置不一致异常"""
    pass


def load_template(
    config: dict,
    template_path: str | os.PathLike[str] | None = None,
) -> Workbook:
    """加载汇总表模板。

    Args:
        config: 由 ConfigLoader.load() 返回的配置字典
        template_path: 可选的运行时模板路径；未提供时使用
            ``runtime.template_path``，相对于 config.yaml 所在目录解析。

    Returns:
        openpyxl Workbook 对象

    Raises:
        FileNotFoundError: 模板文件不存在
    """
    if template_path is None:
        configured_path = config["runtime"]["template_path"]
        config_dir = config["runtime"].get("_config_dir", "")
        # 配置中的模板路径相对于 config.yaml 所在目录。
        full_path = Path(config_dir, configured_path).resolve()
    else:
        if not str(template_path).strip():
            raise ValueError("运行时模板路径不能为空")
        # 命令行指定的路径相对于当前命令执行目录。
        full_path = Path(template_path).expanduser().resolve()

    if not full_path.is_file():
        raise FileNotFoundError(f"模板文件不存在: {full_path}")

    resolved_path = str(full_path)
    logger.info(f"加载模板: {resolved_path}")
    wb = openpyxl.load_workbook(
        resolved_path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )

    # 输出工作表信息
    logger.debug(f"模板包含 {len(wb.sheetnames)} 个工作表: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        merged_count = len(ws.merged_cells.ranges)
        logger.debug(
            f"  {name}: {ws.max_row}行 × {ws.max_column}列, "
            f"合并单元格 {merged_count} 处"
        )

    return wb


def load_data_workbook(file_path: str) -> Workbook:
    """加载权属数据文件。

    Args:
        file_path: 权属文件的完整路径

    Returns:
        openpyxl Workbook 对象

    Raises:
        FileNotFoundError: 文件不存在
    """
    file_path = os.path.abspath(file_path)
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"权属文件不存在: {file_path}")

    wb = openpyxl.load_workbook(
        file_path,
        read_only=False,
        data_only=True,
        keep_links=True,
    )
    return wb


def validate_template(
    wb: Workbook,
    config: dict,
    report_ids: tuple[int, ...] | list[int] | None = None,
) -> list[str]:
    """在写入前校验配置声明与模板结构是否一致。

    对固定行数报表（1/2/3/4/6），检查 B 列值是否与配置的 row_mapping 匹配。
    对报表7，检查每个子表 A 列（园区名称）是否与配置匹配。
    对报表5，检查左右数据区、合计行及写入区合并单元格。
    对报表8，检查数据区边界、列范围及跨表头合并单元格。

    模板中的名称可能包含空格、换行或制表符，因此比对前会移除空白字符；
    归一化后仍采用精确匹配，名称差异应在 row_mapping 中明确配置。

    Args:
        wb: 已加载的模板 Workbook
        config: 配置字典
        report_ids: 仅校验指定报表；None 时保持兼容，校验全部适用报表

    Returns:
        校验通过的报表名称列表

    Raises:
        TemplateMismatchError: 汇总所有不一致后，在模板被修改前抛出
    """
    reports = config["reports"]
    passed = []
    problems = []
    selected_report_ids = (
        set(report_ids) if report_ids is not None else {1, 2, 3, 4, 5, 6, 7, 8}
    )

    for report_id in (1, 2, 3, 4, 6):
        if report_id not in selected_report_ids:
            continue
        report_key = f"report{report_id}"
        rc = reports[report_key]
        sheet_name = rc["sheet_name"]
        if sheet_name not in wb.sheetnames:
            problems.append(f"报表{report_id}: 模板中不存在工作表 '{sheet_name}'")
            continue
        ws = wb[sheet_name]
        b_col = rc["b_col"]
        problem_count_before = len(problems)

        for row_num, expected_name in rc["row_mapping"].items():
            cell_ref = f"{b_col}{row_num}"
            actual_value = ws[cell_ref].value

            # 处理合并单元格中 None 值的情况
            if actual_value is None:
                actual_value = _get_merged_cell_value(ws, cell_ref)

            if actual_value is None:
                problems.append(
                    f"{sheet_name} 第{row_num}行 {b_col}列为空，"
                    f"期望值: '{expected_name}'"
                )
                continue

            if not _names_match(expected_name, str(actual_value)):
                problems.append(
                    f"{sheet_name} 第{row_num}行不匹配: "
                    f"期望 '{expected_name}'，实际 '{str(actual_value).strip()}'"
                )

        if len(problems) == problem_count_before:
            logger.debug(
                f"  {sheet_name}: row_mapping 校验通过 ({len(rc['row_mapping'])}行)"
            )
            passed.append(sheet_name)

    if 5 in selected_report_ids:
        _validate_report5_template(wb, reports["report5"], problems, passed)

    # 校验报表7（园区）
    if 7 in selected_report_ids:
        rc7 = reports["report7"]
        sheet_name = rc7["sheet_name"]
        if sheet_name not in wb.sheetnames:
            problems.append(f"报表7: 模板中不存在工作表 '{sheet_name}'")
        else:
            ws = wb[sheet_name]

            for sub_table in rc7["sub_tables"]:
                sub_name = sub_table["name"]
                problem_count_before = len(problems)
                for row_num, expected_name in sub_table["row_mapping"].items():
                    cell_ref = f"A{row_num}"
                    actual_value = ws[cell_ref].value

                    if actual_value is None:
                        actual_value = _get_merged_cell_value(ws, cell_ref)

                    if actual_value is None:
                        problems.append(
                            f"{sheet_name} [{sub_name}] 第{row_num}行 A列为空，"
                            f"期望值: '{expected_name}'"
                        )
                        continue

                    if not _names_match(expected_name, str(actual_value)):
                        problems.append(
                            f"{sheet_name} [{sub_name}] 第{row_num}行不匹配: "
                            f"期望 '{expected_name}'，实际 '{str(actual_value).strip()}'"
                        )

                if len(problems) == problem_count_before:
                    logger.debug(
                        f"  {sheet_name} [{sub_name}]: row_mapping 校验通过 "
                        f"({len(sub_table['row_mapping'])}个园区)"
                    )
                    passed.append(f"{sheet_name} [{sub_name}]")

    if 8 in selected_report_ids:
        _validate_report8_template(wb, reports["report8"], problems, passed)

    if problems:
        details = "\n".join(
            f"  [{index}] {problem}" for index, problem in enumerate(problems, start=1)
        )
        raise TemplateMismatchError(
            f"模板与配置前置校验失败，共发现 {len(problems)} 项：\n"
            f"{details}\n处理尚未开始，模板未被修改。"
        )

    logger.info(f"模板一致性校验通过: 共 {len(passed)} 个数据区")
    return passed


def _validate_report5_template(
    wb: Workbook,
    report_config: dict,
    problems: list[str],
    passed: list[str],
) -> None:
    """校验报表5动态双栏结构，避免写入合并从属单元格。"""
    sheet_name = report_config["sheet_name"]
    if sheet_name not in wb.sheetnames:
        problems.append(f"报表5: 模板中不存在工作表 '{sheet_name}'")
        return

    worksheet = wb[sheet_name]
    problem_count_before = len(problems)
    sides = {}
    for side_key, side_label in (("left", "左栏"), ("right", "右栏")):
        side_config = report_config.get(side_key)
        path = f"reports.report5.{side_key}"
        if not isinstance(side_config, dict):
            problems.append(f"报表5 {side_label}: {path} 必须是字典")
            continue

        start_row = _positive_row_number(
            side_config.get("data_start_row"), f"{path}.data_start_row", problems
        )
        end_row = _positive_row_number(
            side_config.get("data_end_row"), f"{path}.data_end_row", problems
        )
        total_row = _positive_row_number(
            side_config.get("total_row"), f"{path}.total_row", problems
        )
        columns = _configured_columns(
            side_config.get("cols"), f"{path}.cols", worksheet, problems
        )
        sides[side_key] = (start_row, end_row, total_row, columns)

        if start_row is not None and end_row is not None and start_row > end_row:
            problems.append(
                f"报表5 {side_label}: {path}.data_start_row={start_row} "
                f"不能大于 data_end_row={end_row}"
            )
        if end_row is not None and end_row > worksheet.max_row:
            problems.append(
                f"报表5 {side_label}: {path}.data_end_row={end_row} "
                f"超出模板工作表 '{sheet_name}' 的实际最大行 {worksheet.max_row}"
            )
        if total_row is not None and total_row > worksheet.max_row:
            problems.append(
                f"报表5 {side_label}: {path}.total_row={total_row} "
                f"超出模板工作表 '{sheet_name}' 的实际最大行 {worksheet.max_row}"
            )
        elif total_row is not None and total_row != worksheet.max_row:
            problems.append(
                f"报表5 {side_label}: {path}.total_row={total_row}，"
                f"但模板结构最后一行为第 {worksheet.max_row} 行"
            )
        if end_row is not None and total_row is not None and total_row != end_row + 1:
            problems.append(
                f"报表5 {side_label}: {path}.total_row={total_row} 必须等于 "
                f"data_end_row+1（当前应为 {end_row + 1}）"
            )

        if (
            start_row is not None
            and end_row is not None
            and start_row <= end_row
            and columns
        ):
            for merged_range in worksheet.merged_cells.ranges:
                if _merged_range_intersects_data_area(
                    merged_range, start_row, end_row, columns
                ):
                    problems.append(
                        f"报表5 {side_label}: 合并区域 {merged_range} 落入配置可写数据区 "
                        f"{get_column_letter(columns[0])}{start_row}:"
                        f"{get_column_letter(columns[-1])}{end_row}；写入时将命中只读合并单元格"
                    )

    left = sides.get("left")
    right = sides.get("right")
    if left and right and all(value is not None for value in left[:3] + right[:3]):
        if left[:3] != right[:3]:
            problems.append(
                "报表5: 左右两栏的 data_start_row、data_end_row 和 total_row 必须完全一致，"
                f"当前左栏为 {left[:3]}，右栏为 {right[:3]}"
            )

    if len(problems) == problem_count_before:
        logger.debug(f"  {sheet_name}: 动态双栏结构校验通过")
        passed.append(sheet_name)


def _validate_report8_template(
    wb: Workbook,
    report_config: dict,
    problems: list[str],
    passed: list[str],
) -> None:
    """校验报表8动态明细区边界及合并区域。"""
    sheet_name = report_config["sheet_name"]
    if sheet_name not in wb.sheetnames:
        problems.append(f"报表8: 模板中不存在工作表 '{sheet_name}'")
        return

    worksheet = wb[sheet_name]
    problem_count_before = len(problems)
    start_row = _positive_row_number(
        report_config.get("data_start_row"),
        "reports.report8.data_start_row",
        problems,
    )
    end_row = _positive_row_number(
        report_config.get("data_end_row"),
        "reports.report8.data_end_row",
        problems,
    )
    columns = _configured_columns(
        report_config.get("cols"), "reports.report8.cols", worksheet, problems
    )

    if start_row is not None and end_row is not None and start_row > end_row:
        problems.append(
            f"报表8: reports.report8.data_start_row={start_row} "
            f"不能大于 data_end_row={end_row}"
        )
    if end_row is not None and end_row != worksheet.max_row:
        problems.append(
            f"报表8: reports.report8.data_end_row={end_row} 与模板工作表 "
            f"'{sheet_name}' 的实际最大行 {worksheet.max_row} 不一致"
        )

    if start_row is not None:
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row < start_row <= merged_range.max_row:
                problems.append(
                    f"报表8: 合并区域 {merged_range} 跨越表头和数据区边界 "
                    f"data_start_row={start_row}"
                )
            if (
                end_row is not None
                and merged_range.min_row <= end_row < merged_range.max_row
            ):
                problems.append(
                    f"报表8: 合并区域 {merged_range} 跨越配置数据区结束行 "
                    f"data_end_row={end_row}"
                )
            if (
                columns
                and merged_range.min_row >= start_row
                and (
                    merged_range.min_col < columns[0]
                    or merged_range.max_col > columns[-1]
                )
            ):
                problems.append(
                    f"报表8: 数据区合并区域 {merged_range} 超出配置列范围 "
                    f"{get_column_letter(columns[0])}:{get_column_letter(columns[-1])}"
                )

    if len(problems) == problem_count_before:
        logger.debug(f"  {sheet_name}: 动态明细结构校验通过")
        passed.append(sheet_name)


def _positive_row_number(
    value: object,
    config_path: str,
    problems: list[str],
) -> int | None:
    """读取正整数行号，并将配置问题追加到统一错误列表。"""
    if isinstance(value, bool) or not isinstance(value, int) or value < 1:
        problems.append(f"{config_path} 必须是正整数，实际为 {value!r}")
        return None
    return value


def _configured_columns(
    value: object,
    config_path: str,
    worksheet: Worksheet,
    problems: list[str],
) -> list[int] | None:
    """读取连续且不重复的列配置。"""
    if not isinstance(value, list) or not value:
        problems.append(f"{config_path} 必须是非空列名列表")
        return None

    indices = []
    for column in value:
        if not isinstance(column, str) or not column.strip():
            problems.append(f"{config_path} 包含无效列名 {column!r}")
            return None
        try:
            indices.append(column_index_from_string(column.strip().upper()))
        except ValueError:
            problems.append(f"{config_path} 包含无效列名 {column!r}")
            return None

    expected = list(range(indices[0], indices[-1] + 1))
    if indices != expected:
        problems.append(f"{config_path} 必须按从左到右顺序配置连续且不重复的列")
        return None
    if indices[-1] > worksheet.max_column:
        problems.append(
            f"{config_path} 的结束列 {get_column_letter(indices[-1])} 超出模板工作表 "
            f"'{worksheet.title}' 的实际最大列 {get_column_letter(worksheet.max_column)}"
        )
        return None
    return indices


def _merged_range_intersects_data_area(
    merged_range,
    start_row: int,
    end_row: int,
    columns: list[int],
) -> bool:
    """判断合并区域是否覆盖指定数据行和任一配置列。"""
    if merged_range.max_row < start_row or merged_range.min_row > end_row:
        return False
    return any(
        merged_range.min_col <= column <= merged_range.max_col
        for column in columns
    )


def _normalize_name(name: str) -> str:
    """归一化名称：移除所有空白字符（空格、换行、制表符等）。"""
    return "".join(name.split())


def _names_match(expected: str, actual: str) -> bool:
    """判断两个名称是否匹配（归一化后精确匹配）。

    去除所有空白字符（空格、换行、制表符等）后比对。
    模板中名称有差异的，应在各报表的 row_mapping 中分别配置实际名称。
    """
    return _normalize_name(expected) == _normalize_name(actual)


def _get_merged_cell_value(ws: Worksheet, cell_ref: str) -> str | None:
    """如果单元格属于合并区域且为空，尝试从合并区域左上角获取值。

    Args:
        ws: 工作表
        cell_ref: 单元格引用，如 "B7"

    Returns:
        合并区域左上角的值，如果不在合并区域中则返回 None
    """
    cell = ws[cell_ref]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 获取合并区域左上角单元格的值
            top_left = merged_range.min_row, merged_range.min_col
            top_left_cell = ws.cell(row=top_left[0], column=top_left[1])
            return top_left_cell.value
    return None


def load_ownership_files(config: dict) -> dict:
    """扫描数据目录，按精确文件名加载所有权属文件。

    Args:
        config: 配置字典

    Returns:
        {ownership_key: {"workbook": wb, "file": path, "filename": name}}
        不包含 file 为 null 的条目

    Raises:
        FileNotFoundError: 数据目录不存在，或没有任何可加载的权属文件
    """
    config_dir = config["runtime"].get("_config_dir", "")
    data_dir = os.path.abspath(
        os.path.normpath(
            os.path.join(config_dir, config["runtime"]["data_dir"])
        )
    )

    if not os.path.isdir(data_dir):
        raise FileNotFoundError(f"数据目录不存在: {data_dir}")

    logger.info(f"数据目录: {data_dir}")

    ownership_files = config["ownership_files"]
    expected_filenames = {
        entry["file"]
        for entry in ownership_files.values()
        if entry["file"] is not None
    }
    available_filenames = {
        entry.name
        for entry in os.scandir(data_dir)
        if entry.is_file()
        and entry.name.lower().endswith(".xlsx")
        and not entry.name.startswith("~$")
    }

    for filename in sorted(available_filenames - expected_filenames):
        logger.warning(f"未识别的权属文件，已跳过: {filename}")

    result = {}
    warnings = []
    load_errors = []

    for owner_key, entry in ownership_files.items():
        filename = entry["file"]

        # 无独立数据文件的跳过
        if filename is None:
            logger.debug(f"  跳过 {owner_key}: 无独立数据文件")
            continue

        file_path = os.path.join(data_dir, filename)

        if filename not in available_filenames:
            msg = (
                "未找到权属文件（文件不存在或文件名与配置不一致）："
                f"{file_path}"
            )
            warnings.append(msg)
            continue

        logger.debug(f"  加载 {owner_key}: {filename}")

        try:
            wb = load_data_workbook(file_path)
        except Exception as e:
            msg = f"加载失败 {owner_key} ({filename}): {e}"
            load_errors.append(msg)
            logger.error(msg)
            continue

        result[owner_key] = {
            "workbook": wb,
            "file": file_path,
            "filename": filename,
        }

    for msg in warnings:
        logger.warning(msg)

    if len(result) == 0:
        problem_count = len(warnings) + len(load_errors)
        raise FileNotFoundError(
            f"数据目录中无任何可加载文件，共 {problem_count} 个问题"
        )

    logger.info(f"已加载权属文件: {len(result)} 个")
    return result


def close_ownership_files(ownership_data: dict) -> None:
    """关闭所有权属文件的 workbook。

    Args:
        ownership_data: load_ownership_files() 的返回值
    """
    for owner_key, data in ownership_data.items():
        try:
            data["workbook"].close()
            logger.debug(f"  关闭 {owner_key}")
        except Exception as e:
            logger.warning(f"关闭 {owner_key} 失败: {e}")

"""Report 8 handler: non-owned asset detail appending"""

import logging
from copy import copy

from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from engine.comments import CommentCopyStats, copy_source_comment
from engine.writer import get_merged_ranges, validate_merged_ranges

logger = logging.getLogger(__name__)


def process_report8(
    template_ws: Worksheet,
    ownership_data: dict,
    report_config: dict,
    comment_stats: CommentCopyStats | None = None,
    *,
    excluded_owners: set[str] | None = None,
) -> dict[str, int]:
    """汇总非自有资产明细，并按实际记录数动态调整数据区。"""
    records = _extract_records(
        ownership_data,
        report_config["sheet_name"],
        report_config["data_start_row"],
        report_config["cols"],
        excluded_owners=excluded_owners,
    )
    header_merges = _unmerge_data_ranges(
        template_ws,
        report_config["data_start_row"],
    )
    data_end_row, expanded_rows, removed_rows = _resize_data_area(
        template_ws,
        report_config,
        len(records),
    )
    _clear_data_area(
        template_ws,
        report_config["data_start_row"],
        data_end_row,
        report_config["cols"],
    )
    target_rows = _write_records(
        template_ws,
        records,
        report_config["data_start_row"],
        report_config["cols"],
        comment_stats=comment_stats,
    )
    data_merges = _recreate_data_merges(
        template_ws,
        records,
        target_rows,
        report_config["data_start_row"],
        report_config["cols"],
    )
    validate_merged_ranges(
        template_ws,
        frozenset(set(header_merges) | data_merges),
    )

    logger.info(
        f"报表8汇总完成: {len(records)} 条，数据区至第 {data_end_row} 行"
    )
    return {
        "record_count": len(records),
        "expanded_rows": expanded_rows,
        "removed_rows": removed_rows,
        "data_end_row": data_end_row,
    }


def _extract_records(
    ownership_data: dict,
    sheet_name: str,
    data_start_row: int,
    columns: list[str],
    *,
    excluded_owners: set[str] | None = None,
) -> list[dict]:
    """提取所有权属的有效明细，并保留来源行和工作表。"""
    column_indices = [column_index_from_string(column) for column in columns]
    records = []
    excluded_owners = excluded_owners or set()

    for owner_key, owner_data in ownership_data.items():
        if owner_key in excluded_owners:
            continue
        workbook = owner_data["workbook"]
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"报表8: 权属 '{owner_key}' 缺少工作表 '{sheet_name}'，跳过")
            continue

        worksheet = workbook[sheet_name]
        owner_count = 0
        for row in range(data_start_row, worksheet.max_row + 1):
            values = [worksheet.cell(row=row, column=column).value for column in column_indices]
            if not any(_has_content(value) for value in values[1:]):
                continue

            records.append(
                {
                    "owner": owner_key,
                    "worksheet": worksheet,
                    "source_row": row,
                }
            )
            owner_count += 1

        logger.debug(f"  报表8 {owner_key}: {owner_count} 条")

    return records


def _unmerge_data_ranges(
    worksheet: Worksheet,
    data_start_row: int,
) -> frozenset[str]:
    """解除旧明细绑定的合并区域，保留表头合并区域。"""
    all_ranges = list(worksheet.merged_cells.ranges)
    header_ranges = set()
    data_ranges = []

    for merged_range in all_ranges:
        if merged_range.max_row < data_start_row:
            header_ranges.add(str(merged_range))
        elif merged_range.min_row >= data_start_row:
            data_ranges.append(merged_range)
        else:
            raise ValueError(
                f"报表8存在跨越表头和数据区的合并区域: {merged_range}"
            )

    for merged_range in data_ranges:
        worksheet.unmerge_cells(str(merged_range))

    if data_ranges:
        logger.debug(f"报表8解除 {len(data_ranges)} 个旧数据区合并区域")
    return frozenset(header_ranges)


def _resize_data_area(
    worksheet: Worksheet,
    report_config: dict,
    required_rows: int,
) -> tuple[int, int, int]:
    """按总明细数扩展或收缩单栏数据区。"""
    data_start_row = report_config["data_start_row"]
    data_end_row = report_config["data_end_row"]
    capacity = data_end_row - data_start_row + 1
    row_delta = required_rows - capacity

    if row_delta > 0:
        worksheet.insert_rows(data_end_row + 1, amount=row_delta)
        _copy_data_row_style(
            worksheet,
            source_row=data_end_row,
            start_row=data_end_row + 1,
            end_row=data_end_row + row_delta,
            start_col=1,
            end_col=len(report_config["cols"]),
        )
        new_data_end_row = data_end_row + row_delta
        logger.warning(
            f"报表8明细超出模板容量 {row_delta} 行，"
            f"数据区已扩展至第 {new_data_end_row} 行"
        )
        return new_data_end_row, row_delta, 0

    if row_delta < 0:
        removed_rows = -row_delta
        delete_start_row = data_start_row + required_rows
        worksheet.delete_rows(delete_start_row, amount=removed_rows)
        new_data_end_row = data_end_row - removed_rows
        logger.info(
            f"报表8明细少于模板容量 {removed_rows} 行，"
            f"数据区已收缩至第 {new_data_end_row} 行"
        )
        return new_data_end_row, 0, removed_rows

    return data_end_row, 0, 0


def _copy_data_row_style(
    worksheet: Worksheet,
    source_row: int,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """将最后一个预置数据行的样式复制给新增行。"""
    source_height = worksheet.row_dimensions[source_row].height
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = source_height
        for column in range(start_col, end_col + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=row, column=column)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)


def _clear_data_area(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    columns: list[str],
) -> None:
    """清空数据区的值，保留单元格样式。"""
    column_indices = [column_index_from_string(column) for column in columns]
    for row in range(start_row, end_row + 1):
        for column in column_indices:
            worksheet.cell(row=row, column=column).value = None


def _write_records(
    worksheet: Worksheet,
    records: list[dict],
    data_start_row: int,
    columns: list[str],
    *,
    comment_stats: CommentCopyStats | None = None,
) -> dict[tuple[str, int], int]:
    """完整复制来源行数据和样式，并从1开始重编序号。"""
    column_indices = [column_index_from_string(column) for column in columns]
    target_rows = {}
    for sequence, record in enumerate(records, start=1):
        row = data_start_row + sequence - 1
        source_ws = record["worksheet"]
        source_row = record["source_row"]
        worksheet.row_dimensions[row].height = source_ws.row_dimensions[source_row].height
        for column in column_indices:
            source_cell = source_ws.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=row, column=column)
            _copy_cell_style(source_cell, target_cell)
            target_cell.value = source_cell.value
            if column != column_indices[0]:
                copy_source_comment(
                    source_cell,
                    target_cell,
                    report_id=8,
                    owner_key=record["owner"],
                    source_ws=source_ws,
                    target_ws=worksheet,
                    stats=comment_stats,
                )

        worksheet.cell(row=row, column=column_indices[0]).value = sequence
        target_rows[(record["owner"], source_row)] = row
    return target_rows


def _copy_cell_style(source_cell, target_cell) -> None:
    """跨工作簿复制样式，并在目标工作簿中注册各样式组件。"""
    if not source_cell.has_style:
        return

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _recreate_data_merges(
    target_ws: Worksheet,
    records: list[dict],
    target_rows: dict[tuple[str, int], int],
    data_start_row: int,
    columns: list[str],
) -> set[str]:
    """按来源数据行的相对位置，在汇总表重建有效合并区域。"""
    min_col = column_index_from_string(columns[0])
    max_col = column_index_from_string(columns[-1])
    source_sheets = {}
    for record in records:
        source_sheets[record["owner"]] = record["worksheet"]

    recreated = set()
    for owner_key, source_ws in source_sheets.items():
        for merged_range in source_ws.merged_cells.ranges:
            if (
                merged_range.min_row < data_start_row
                or merged_range.min_col < min_col
                or merged_range.max_col > max_col
            ):
                continue

            source_rows = list(range(merged_range.min_row, merged_range.max_row + 1))
            if not all((owner_key, row) in target_rows for row in source_rows):
                continue

            mapped_rows = [target_rows[(owner_key, row)] for row in source_rows]
            if mapped_rows != list(range(mapped_rows[0], mapped_rows[-1] + 1)):
                continue

            target_range = (
                f"{get_column_letter(merged_range.min_col)}"
                f"{mapped_rows[0]}:"
                f"{get_column_letter(merged_range.max_col)}"
                f"{mapped_rows[-1]}"
            )
            target_ws.merge_cells(target_range)
            recreated.add(target_range)

    return recreated


def _has_content(value: object) -> bool:
    """判断明细字段是否有内容；数字0视为有效。"""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True

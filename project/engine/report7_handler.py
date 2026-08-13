"""Report 7 handler: multiple sub-tables with园区 matching"""

import logging

from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from engine.comments import CommentCopyStats, copy_source_comment
from engine.matcher import find_unit_row
from engine.writer import (
    _is_numeric_value,
    get_merged_ranges,
    validate_merged_ranges,
)

logger = logging.getLogger(__name__)


class ParkSourceError(Exception):
    """园区数据来源缺失或不唯一。"""


def process_report7(
    template_ws: Worksheet,
    ownership_data: dict,
    report_config: dict,
    comment_stats: CommentCopyStats | None = None,
    *,
    excluded_owners: set[str] | None = None,
) -> dict[str, dict]:
    """按园区名称从权属文件匹配数据并写入四个子表。"""
    merged_ranges_before = get_merged_ranges(template_ws)
    sheet_name = report_config["sheet_name"]
    start_col = column_index_from_string(report_config["data_start_col"])
    end_col = column_index_from_string(report_config["data_end_col"])
    formula_columns = {
        column_index_from_string(str(column).upper())
        for column in report_config.get("formula_columns", [])
    }
    text_columns = {
        column_index_from_string(str(column).upper())
        for column in report_config.get("text_columns", [])
    }
    results = {}
    excluded_owners = excluded_owners or set()

    for sub_table in report_config["sub_tables"]:
        for target_row, park_name in sub_table["row_mapping"].items():
            owner_key, source_ws, source_row = _find_unique_park_source(
                ownership_data,
                sheet_name,
                park_name,
                excluded_owners=excluded_owners,
            )
            written_cells = _write_park_row(
                template_ws,
                target_row,
                source_ws,
                source_row,
                start_col,
                end_col,
                formula_columns=formula_columns,
                text_columns=text_columns,
                owner_key=owner_key,
                comment_stats=comment_stats,
            )
            results[park_name] = {
                "owner": owner_key,
                "source_row": source_row,
                "target_row": target_row,
                "written_cells": written_cells,
            }
            logger.debug(
                f"  报表7 {park_name}: {owner_key} 第{source_row}行 "
                f"→ 汇总第{target_row}行，写入 {written_cells} 个单元格"
            )

    validate_merged_ranges(template_ws, merged_ranges_before)
    logger.info(f"报表7汇总完成: {len(results)} 个园区")
    return results


def _find_unique_park_source(
    ownership_data: dict,
    sheet_name: str,
    park_name: str,
    *,
    excluded_owners: set[str] | None = None,
) -> tuple[str, Worksheet, int]:
    """查找唯一具有实际运营数据的园区来源行。"""
    candidates = []
    name_matches = []
    excluded_owners = excluded_owners or set()

    for owner_key, owner_data in ownership_data.items():
        if owner_key in excluded_owners:
            continue
        workbook = owner_data["workbook"]
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        source_row = find_unit_row(
            worksheet,
            park_name,
            col_letter="A",
            data_start_row=1,
        )
        if source_row is None:
            continue

        name_matches.append(owner_key)
        if _has_operating_data(worksheet, source_row):
            candidates.append((owner_key, worksheet, source_row))

    if len(candidates) == 1:
        return candidates[0]

    if not candidates:
        matched_text = ", ".join(name_matches) if name_matches else "无"
        raise ParkSourceError(
            f"园区 '{park_name}' 未找到唯一有效数据来源；"
            f"名称出现于: {matched_text}"
        )

    owners = ", ".join(candidate[0] for candidate in candidates)
    raise ParkSourceError(
        f"园区 '{park_name}' 存在多个有效数据来源: {owners}"
    )


def _has_operating_data(worksheet: Worksheet, row: int) -> bool:
    """D:T 任一单元格有内容即视为该园区已实际填报。"""
    start_col = column_index_from_string("D")
    end_col = column_index_from_string("T")
    return any(
        _has_content(worksheet.cell(row=row, column=column).value)
        for column in range(start_col, end_col + 1)
    )


def _write_park_row(
    target_ws: Worksheet,
    target_row: int,
    source_ws: Worksheet,
    source_row: int,
    start_col: int,
    end_col: int,
    *,
    formula_columns: set[int] | None = None,
    text_columns: set[int] | None = None,
    owner_key: str,
    comment_stats: CommentCopyStats | None = None,
) -> int:
    """复制园区明细行，并保留配置公式列中的模板公式。

    A 列园区名称保持模板值。B:V 中 ``formula_columns`` 跳过写入，
    ``text_columns`` 保留来源原值；其他列仅保留数值，空值、``/`` 和
    其他非数值统一写入 0。数值列即使目标原值是公式，也仍会被覆盖。
    """
    formula_columns = formula_columns or set()
    text_columns = text_columns or set()
    written_cells = 0
    for column in range(start_col + 1, end_col + 1):
        if column in formula_columns:
            continue
        target_cell = target_ws.cell(row=target_row, column=column)
        if isinstance(target_cell, MergedCell):
            continue
        source_cell = source_ws.cell(row=source_row, column=column)
        source_value = source_cell.value
        if column in text_columns:
            target_cell.value = source_value
        elif _is_numeric_value(source_value):
            target_cell.value = source_value
        else:
            target_cell.value = 0
        copy_source_comment(
            source_cell,
            target_cell,
            report_id=7,
            owner_key=owner_key,
            source_ws=source_ws,
            target_ws=target_ws,
            stats=comment_stats,
        )
        written_cells += 1
    return written_cells


def _has_content(value: object) -> bool:
    """判断单元格是否包含实际内容，数字0视为有效。"""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True

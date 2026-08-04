"""Generate the next-quarter input template from an approved summary workbook."""

from __future__ import annotations

from copy import copy, deepcopy
from dataclasses import dataclass
import logging
from pathlib import Path
import re
from typing import Iterable

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from engine.comments import clear_template_comments
from engine.output import (
    _build_output_filename,
    _resolve_output_dir,
    save_summary_workbook,
)
from engine.period import QuarterContext


logger = logging.getLogger(__name__)
_TOTAL_LABELS = {"合计", "汇总", "总计"}


class NextTemplateError(Exception):
    """The next-quarter template could not be generated safely."""


@dataclass(frozen=True)
class NextTemplateConsistencyIssue:
    """One source-workbook difference that requires manual review."""

    report_key: str
    sheet_name: str
    category: str
    detail: str

    def format(self) -> str:
        report_label = self.report_key.replace("report", "报表")
        return (
            f"[配置一致性检查] {report_label}｜{self.sheet_name}｜"
            f"{self.category}：{self.detail}"
        )


class NextTemplateConsistencyError(NextTemplateError):
    """The source workbook and template configuration need manual review."""

    def __init__(self, issues: Iterable[NextTemplateConsistencyIssue]):
        self.issues = tuple(issues)
        super().__init__(
            f"源汇总表与配置不一致，共检测到 {len(self.issues)} 项问题；"
            "未生成下季度模板，请人工核对配置和源表后重试"
        )


@dataclass(frozen=True)
class NextTemplateResult:
    """Summary of one successful next-template generation."""

    output_path: str
    source_path: str
    source_quarter: str
    target_quarter: str
    header_replacements: int
    rolled_cells: int
    cleared_cells: int
    comments_cleared: int


def generate_next_template(
    source_path: str | Path,
    current_period: QuarterContext,
    config: dict,
) -> NextTemplateResult:
    """Create and save the immediately following quarter's template."""
    resolved_source = Path(source_path).expanduser().resolve()
    if not resolved_source.is_file():
        raise FileNotFoundError(f"当前季度最终汇总表不存在: {resolved_source}")
    if resolved_source.suffix.lower() != ".xlsx":
        raise NextTemplateError("当前季度最终汇总表必须是 .xlsx 文件")

    rules = _load_rules(config)
    target_period = current_period.next()
    formula_workbook: Workbook | None = None
    values_workbook: Workbook | None = None

    try:
        formula_workbook = openpyxl.load_workbook(
            resolved_source,
            read_only=False,
            data_only=False,
            keep_links=True,
        )
        values_workbook = openpyxl.load_workbook(
            resolved_source,
            read_only=False,
            data_only=True,
            keep_links=True,
        )
        _validate_report_sheets(formula_workbook, values_workbook, config, rules)
        structure_before = _capture_structure(formula_workbook)

        marker_count = _count_source_period_markers(
            formula_workbook,
            current_period,
            config,
            rules,
        )
        if marker_count == 0:
            raise NextTemplateError(
                f"未在配置的标题区域识别到源季度 {current_period.label} "
                f"或截止日期 {current_period.end_date}；请核对 --quarter 与源文件"
            )

        _validate_source_config_consistency(formula_workbook, config, rules)

        header_replacements = _replace_period_headers(
            formula_workbook,
            current_period,
            target_period,
            config,
            rules,
        )
        cleared_cells = 0
        rolled_cells = 0
        for report_key, rule in rules["reports"].items():
            report_config = config["reports"][report_key]
            worksheet = formula_workbook[report_config["sheet_name"]]
            values_worksheet = values_workbook[report_config["sheet_name"]]
            cleared, rolled = _process_report(
                report_key,
                worksheet,
                values_worksheet,
                report_config,
                rule,
            )
            cleared_cells += cleared
            rolled_cells += rolled

        comments_cleared = clear_template_comments(formula_workbook)
        _validate_structure_unchanged(formula_workbook, structure_before)
        fills_cleared, rows_removed = _sanitize_template(
            formula_workbook,
            config,
            rules,
        )
        _enable_excel_recalculation(formula_workbook)
        _validate_no_comments(formula_workbook)
        _validate_sanitized_template(formula_workbook, config, rules)

        save_config = deepcopy(config)
        save_config["quarter"] = target_period.as_config()
        save_config["runtime"] = dict(config["runtime"])
        save_config["runtime"]["output_filename"] = rules["output_filename"]
        target_path = _resolve_output_dir(save_config) / _build_output_filename(
            save_config
        )
        if target_path.resolve() == resolved_source:
            raise NextTemplateError("下季度模板输出路径不能与源汇总表相同")

        output_path = save_summary_workbook(
            formula_workbook,
            save_config,
            artifact_label="下季度模板",
        )
        logger.info(
            "下季度模板生成完成: %s → %s；表头替换 %s 项，滚转 %s 个单元格，"
            "清空 %s 个单元格，清除批注 %s 条，清除遗留填充 %s 处，"
            "删除明细行 %s 行",
            current_period.label,
            target_period.label,
            header_replacements,
            rolled_cells,
            cleared_cells,
            comments_cleared,
            fills_cleared,
            rows_removed,
        )
        return NextTemplateResult(
            output_path=output_path,
            source_path=str(resolved_source),
            source_quarter=current_period.code,
            target_quarter=target_period.code,
            header_replacements=header_replacements,
            rolled_cells=rolled_cells,
            cleared_cells=cleared_cells,
            comments_cleared=comments_cleared,
        )
    finally:
        if values_workbook is not None:
            values_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()


def _load_rules(config: dict) -> dict:
    rules = config.get("next_template")
    if not isinstance(rules, dict):
        raise NextTemplateError("配置文件缺少 next_template 节")
    output_filename = rules.get("output_filename")
    report_rules = rules.get("reports")
    if not isinstance(output_filename, str) or not output_filename.strip():
        raise NextTemplateError("next_template.output_filename 必须是非空字符串")
    if not isinstance(report_rules, dict):
        raise NextTemplateError("next_template.reports 必须是字典")
    for report_id in range(1, 9):
        key = f"report{report_id}"
        if key not in report_rules:
            raise NextTemplateError(f"next_template.reports 缺少 {key}")
    return rules


def _validate_report_sheets(
    formula_workbook: Workbook,
    values_workbook: Workbook,
    config: dict,
    rules: dict,
) -> None:
    for report_key in rules["reports"]:
        try:
            sheet_name = config["reports"][report_key]["sheet_name"]
        except (KeyError, TypeError) as exc:
            raise NextTemplateError(f"缺少 {report_key} 的报表配置") from exc
        if sheet_name not in formula_workbook.sheetnames:
            raise NextTemplateError(f"源汇总表缺少工作表: {sheet_name}")
        if sheet_name not in values_workbook.sheetnames:
            raise NextTemplateError(f"源汇总表数值视图缺少工作表: {sheet_name}")


def _validate_source_config_consistency(
    workbook: Workbook,
    config: dict,
    rules: dict,
) -> None:
    """Detect source/config drift before the workbook is modified."""
    issues: list[NextTemplateConsistencyIssue] = []
    checked = 0
    for report_key, rule in rules["reports"].items():
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        kind = rule.get("kind")
        if kind == "fixed":
            checked += 1
            issues.extend(
                _detect_mapped_region_issues(
                    report_key,
                    worksheet,
                    report_config["row_mapping"],
                    report_config["data_start_row"],
                    report_config["data_end_row"],
                    report_config["b_col"],
                    configured_total_row=report_config.get("total_row"),
                    total_columns=(
                        report_config.get("a_col", "A"),
                        report_config["b_col"],
                    ),
                )
            )
        elif kind == "double_detail":
            checked += 1
            issues.extend(
                _detect_double_detail_consistency_issues(
                    report_key,
                    worksheet,
                    report_config,
                )
            )
        elif kind == "park":
            checked += 1
            issues.extend(
                _detect_park_consistency_issues(
                    report_key,
                    worksheet,
                    report_config,
                )
            )

    if not issues:
        logger.info("源汇总表与配置一致性检查通过: 共检查 %s 个报表", checked)
        return

    logger.warning("源汇总表与配置不一致，需人工核对以下问题:")
    for issue in issues:
        logger.warning("  %s", issue.format())
    raise NextTemplateConsistencyError(issues)


def _detect_mapped_region_issues(
    report_key: str,
    worksheet: Worksheet,
    row_mapping: dict,
    data_start_row: int,
    configured_data_end_row: int,
    unit_column: str,
    *,
    configured_total_row: int | None,
    total_columns: tuple[str, ...],
    context: str | None = None,
    search_end_row: int | None = None,
) -> list[NextTemplateConsistencyIssue]:
    issues: list[NextTemplateConsistencyIssue] = []
    actual_total_row = _find_first_total_row(
        worksheet,
        data_start_row,
        total_columns,
        search_end_row,
    )
    context_prefix = f"{context}，" if context else ""

    if isinstance(configured_total_row, int):
        if actual_total_row is None:
            issues.append(
                _consistency_issue(
                    report_key,
                    worksheet,
                    "合计行缺失",
                    f"{context_prefix}配置第{configured_total_row}行，来源表未找到合计标识",
                )
            )
        elif actual_total_row != configured_total_row:
            issues.append(
                _consistency_issue(
                    report_key,
                    worksheet,
                    "合计行变化",
                    f"{context_prefix}配置第{configured_total_row}行，实际第{actual_total_row}行",
                )
            )
    elif actual_total_row is not None:
        issues.append(
            _consistency_issue(
                report_key,
                worksheet,
                "出现未配置合计行",
                f"{context_prefix}实际第{actual_total_row}行存在合计标识",
            )
        )

    scan_end_row = (
        actual_total_row - 1
        if actual_total_row is not None
        else search_end_row or configured_data_end_row
    )
    actual_entries = _mapped_entries(
        worksheet,
        data_start_row,
        scan_end_row,
        unit_column,
    )
    actual_end_row = max(
        (row for row, _, _ in actual_entries),
        default=data_start_row - 1,
    )
    if actual_end_row != configured_data_end_row:
        issues.append(
            _consistency_issue(
                report_key,
                worksheet,
                "数据区结束行变化",
                f"{context_prefix}配置第{configured_data_end_row}行，实际第{actual_end_row}行",
            )
        )

    expected_by_name: dict[str, list[tuple[int, str]]] = {}
    for row, name in sorted(row_mapping.items()):
        normalized = _normalize_consistency_name(name)
        expected_by_name.setdefault(normalized, []).append((row, str(name)))

    actual_by_name: dict[str, list[tuple[int, str]]] = {}
    for row, raw_name, normalized in actual_entries:
        actual_by_name.setdefault(normalized, []).append((row, raw_name))

    for normalized, entries in actual_by_name.items():
        if len(entries) > 1:
            detail = "、".join(f"第{row}行“{name}”" for row, name in entries)
            issues.append(
                _consistency_issue(
                    report_key,
                    worksheet,
                    "重复单位",
                    f"{context_prefix}{detail}",
                )
            )
        if normalized not in expected_by_name:
            for row, name in entries:
                issues.append(
                    _consistency_issue(
                        report_key,
                        worksheet,
                        "新增单位",
                        f"{context_prefix}来源第{row}行“{name}”未出现在 row_mapping",
                    )
                )

    for normalized, expected_entries in expected_by_name.items():
        actual_matches = actual_by_name.get(normalized, [])
        if not actual_matches:
            for row, name in expected_entries:
                issues.append(
                    _consistency_issue(
                        report_key,
                        worksheet,
                        "配置单位缺失",
                        f"{context_prefix}配置第{row}行“{name}”未在来源数据区找到",
                    )
                )
            continue
        if len(expected_entries) == 1 and len(actual_matches) == 1:
            expected_row, expected_name = expected_entries[0]
            actual_row, _ = actual_matches[0]
            if actual_row != expected_row:
                issues.append(
                    _consistency_issue(
                        report_key,
                        worksheet,
                        "单位行位移",
                        f"{context_prefix}“{expected_name}”配置第{expected_row}行，实际第{actual_row}行",
                    )
                )
    return issues


def _detect_double_detail_consistency_issues(
    report_key: str,
    worksheet: Worksheet,
    report_config: dict,
) -> list[NextTemplateConsistencyIssue]:
    totals: dict[str, int | None] = {}
    issues: list[NextTemplateConsistencyIssue] = []
    for side_name in ("left", "right"):
        side = report_config[side_name]
        total_row = _find_last_total_row(
            worksheet,
            side["data_start_row"],
            side["cols"][0],
        )
        totals[side_name] = total_row
        if total_row is None:
            issues.append(
                _consistency_issue(
                    report_key,
                    worksheet,
                    "合计行缺失",
                    f"{side_name} 数据区未找到合计标识",
                )
            )
    if all(total is not None for total in totals.values()) and (
        totals["left"] != totals["right"]
    ):
        issues.append(
            _consistency_issue(
                report_key,
                worksheet,
                "左右合计行不一致",
                f"左侧第{totals['left']}行，右侧第{totals['right']}行",
            )
        )
    return issues


def _detect_park_consistency_issues(
    report_key: str,
    worksheet: Worksheet,
    report_config: dict,
) -> list[NextTemplateConsistencyIssue]:
    issues: list[NextTemplateConsistencyIssue] = []
    sub_tables = report_config["sub_tables"]
    for index, sub_table in enumerate(sub_tables):
        next_header_row = (
            sub_tables[index + 1]["header_row"]
            if index + 1 < len(sub_tables)
            else None
        )
        search_end_row = (
            next_header_row - 2
            if next_header_row is not None
            else worksheet.max_row
        )
        issues.extend(
            _detect_mapped_region_issues(
                report_key,
                worksheet,
                sub_table["row_mapping"],
                sub_table["data_start_row"],
                sub_table["data_end_row"],
                "A",
                configured_total_row=sub_table.get("total_row"),
                total_columns=("A",),
                context=f"子表“{sub_table['name']}”",
                search_end_row=search_end_row,
            )
        )
    return issues


def _mapped_entries(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    unit_column: str,
) -> list[tuple[int, str, str]]:
    entries: list[tuple[int, str, str]] = []
    for row in range(start_row, end_row + 1):
        value = _cell_or_merged_value(worksheet, f"{unit_column}{row}")
        normalized = _normalize_consistency_name(value)
        if normalized and not _is_total_label(value):
            entries.append((row, str(value).strip(), normalized))
    return entries


def _find_first_total_row(
    worksheet: Worksheet,
    start_row: int,
    columns: tuple[str, ...],
    end_row: int | None = None,
) -> int | None:
    final_row = min(end_row or worksheet.max_row, worksheet.max_row)
    for row in range(start_row, final_row + 1):
        if any(
            _is_total_label(_cell_or_merged_value(worksheet, f"{column}{row}"))
            for column in columns
        ):
            return row
    return None


def _find_last_total_row(
    worksheet: Worksheet,
    start_row: int,
    column: str,
) -> int | None:
    for row in range(worksheet.max_row, start_row - 1, -1):
        if _is_total_label(_cell_or_merged_value(worksheet, f"{column}{row}")):
            return row
    return None


def _cell_or_merged_value(worksheet: Worksheet, cell_ref: str) -> object | None:
    cell = worksheet[cell_ref]
    if cell.value is not None:
        return cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
    return None


def _normalize_consistency_name(value: object | None) -> str:
    return "" if value is None else "".join(str(value).split())


def _is_total_label(value: object | None) -> bool:
    normalized = _normalize_consistency_name(value)
    return normalized in _TOTAL_LABELS or any(
        normalized.endswith(label) for label in _TOTAL_LABELS
    )


def _consistency_issue(
    report_key: str,
    worksheet: Worksheet,
    category: str,
    detail: str,
) -> NextTemplateConsistencyIssue:
    return NextTemplateConsistencyIssue(
        report_key=report_key,
        sheet_name=worksheet.title,
        category=category,
        detail=detail,
    )


def _process_report(
    report_key: str,
    worksheet: Worksheet,
    values_worksheet: Worksheet,
    report_config: dict,
    rule: dict,
) -> tuple[int, int]:
    kind = rule.get("kind")
    if kind == "fixed":
        return _process_fixed_report(
            worksheet,
            values_worksheet,
            report_config,
            rule,
        )
    if kind == "double_detail":
        return _process_double_detail(worksheet, report_config), 0
    if kind == "park":
        return _process_park_report(worksheet, report_config, rule), 0
    if kind == "detail":
        return _process_detail_report(worksheet, report_config), 0
    raise NextTemplateError(f"{report_key} 的 next_template.kind 不受支持: {kind!r}")


def _process_fixed_report(
    worksheet: Worksheet,
    values_worksheet: Worksheet,
    report_config: dict,
    rule: dict,
) -> tuple[int, int]:
    rows = range(report_config["data_start_row"], report_config["data_end_row"] + 1)
    clear_columns = _expand_column_specs(rule.get("clear_columns", []))
    formula_columns = _expand_column_specs(
        rule.get("preserve_formula_columns", [])
    )
    cleared = _clear_cells(
        worksheet,
        rows,
        clear_columns,
        preserve_formula_columns=formula_columns,
    )

    rolled = 0
    roll_columns = rule.get("roll_columns", {})
    if not isinstance(roll_columns, dict):
        raise NextTemplateError("roll_columns 必须是字典")
    for source_column, target_column in roll_columns.items():
        for row in rows:
            source_cell = worksheet[f"{source_column}{row}"]
            target_cell = worksheet[f"{target_column}{row}"]
            value = values_worksheet[f"{source_column}{row}"].value
            if isinstance(target_cell, MergedCell):
                if value is None:
                    continue
                raise NextTemplateError(
                    f"滚转目标是合并区域的非左上角单元格，且源值非空: "
                    f"{worksheet.title}!{target_column}{row}"
                )
            if not isinstance(source_cell, MergedCell) and (
                source_cell.data_type == "f" and value is None
            ):
                raise NextTemplateError(
                    f"滚转源公式没有已保存的计算结果: "
                    f"{worksheet.title}!{source_column}{row}"
                )
            target_cell.value = value
            rolled += 1

    total_row = report_config.get("total_row")
    if isinstance(total_row, int):
        total_columns = _columns_between(
            report_config["data_start_col"],
            report_config["data_end_col"],
        )
        cleared += _clear_cells(
            worksheet,
            [total_row],
            total_columns,
            preserve_all_formulas=True,
        )
    return cleared, rolled


def _process_double_detail(worksheet: Worksheet, report_config: dict) -> int:
    cleared = 0
    for side_name in ("left", "right"):
        side = report_config[side_name]
        columns = side["cols"]
        sequence_column = columns[0]
        total_row = _find_total_row(
            worksheet,
            side["data_start_row"],
            sequence_column,
            side.get("total_row"),
        )
        data_columns = columns[1:]
        for sequence, row in enumerate(
            range(side["data_start_row"], total_row),
            start=1,
        ):
            sequence_cell = worksheet[f"{sequence_column}{row}"]
            if not isinstance(sequence_cell, MergedCell):
                sequence_cell.value = sequence
            cleared += _clear_cells(worksheet, [row], data_columns)
    return cleared


def _process_park_report(
    worksheet: Worksheet,
    report_config: dict,
    rule: dict,
) -> int:
    clear_columns = _expand_column_specs(rule.get("clear_columns", []))
    formula_columns = _expand_column_specs(
        rule.get("preserve_formula_columns", [])
    )
    cleared = 0
    for sub_table in report_config["sub_tables"]:
        rows = range(sub_table["data_start_row"], sub_table["data_end_row"] + 1)
        cleared += _clear_cells(
            worksheet,
            rows,
            clear_columns,
            preserve_formula_columns=formula_columns,
        )
        total_row = sub_table.get("total_row")
        if isinstance(total_row, int):
            cleared += _clear_cells(
                worksheet,
                [total_row],
                clear_columns,
                preserve_all_formulas=True,
            )
    return cleared


def _process_detail_report(worksheet: Worksheet, report_config: dict) -> int:
    columns = report_config["cols"]
    sequence_column = columns[0]
    data_columns = columns[1:]
    start_row = report_config["data_start_row"]
    end_row = max(report_config["data_end_row"], worksheet.max_row)
    cleared = 0
    for sequence, row in enumerate(range(start_row, end_row + 1), start=1):
        sequence_cell = worksheet[f"{sequence_column}{row}"]
        if not isinstance(sequence_cell, MergedCell):
            sequence_cell.value = sequence
        cleared += _clear_cells(worksheet, [row], data_columns)
    return cleared


def _sanitize_template(
    workbook: Workbook,
    config: dict,
    rules: dict,
) -> tuple[int, int]:
    """Remove summary-only presentation and dynamic-detail residue."""
    rows_removed = 0

    for report_key, rule in rules["reports"].items():
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        target_max_row = rule.get("target_max_row")

        if target_max_row is not None and (
            not isinstance(target_max_row, int) or target_max_row < 1
        ):
            raise NextTemplateError(
                f"next_template.reports.{report_key}.target_max_row 必须是正整数"
            )

        kind = rule.get("kind")
        if kind == "double_detail":
            _, current_total_row = _double_detail_bounds(
                worksheet,
                report_config,
            )
            if target_max_row is not None:
                rows_removed += _normalize_double_detail_rows(
                    worksheet,
                    report_config,
                    target_max_row,
                    current_total_row,
                )
        elif kind == "detail":
            if target_max_row is not None:
                rows_removed += _normalize_detail_rows(
                    worksheet,
                    report_config,
                    target_max_row,
                )
        elif kind == "fixed":
            if target_max_row is not None:
                raise NextTemplateError(
                    f"{report_key} 的 target_max_row 仅支持明细类报表"
                )
        elif kind == "park":
            if target_max_row is not None:
                raise NextTemplateError(
                    f"{report_key} 的 target_max_row 仅支持明细类报表"
                )
        elif target_max_row is not None:
            raise NextTemplateError(
                f"{report_key} 的 target_max_row 仅支持明细类报表"
            )

    clear_manual_fills = rules.get("clear_manual_fills", False)
    if not isinstance(clear_manual_fills, bool):
        raise NextTemplateError("next_template.clear_manual_fills 必须是布尔值")
    fills_cleared = (
        _clear_manual_data_fills(workbook, config, rules)
        if clear_manual_fills
        else 0
    )
    return fills_cleared, rows_removed


def _double_detail_bounds(
    worksheet: Worksheet,
    report_config: dict,
) -> tuple[int, int]:
    left = report_config["left"]
    right = report_config["right"]
    if left["data_start_row"] != right["data_start_row"]:
        raise NextTemplateError(f"{worksheet.title} 左右数据区起始行不一致")

    totals = {
        _find_total_row(
            worksheet,
            side["data_start_row"],
            side["cols"][0],
            side.get("total_row"),
        )
        for side in (left, right)
    }
    if len(totals) != 1:
        raise NextTemplateError(f"{worksheet.title} 左右合计行不一致")
    return left["data_start_row"], totals.pop()


def _normalize_double_detail_rows(
    worksheet: Worksheet,
    report_config: dict,
    target_total_row: int,
    current_total_row: int,
) -> int:
    """Move a double-detail report's total row to the configured final row."""
    data_start_row = report_config["left"]["data_start_row"]
    if target_total_row <= data_start_row:
        raise NextTemplateError(
            f"{worksheet.title} 目标总行数必须大于数据起始行 {data_start_row}"
        )

    total_merges = [
        merged_range
        for merged_range in list(worksheet.merged_cells.ranges)
        if merged_range.min_row <= current_total_row <= merged_range.max_row
    ]
    total_merge_bounds = [
        (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        for merged_range in total_merges
    ]
    formula_columns = (
        report_config["left"]["cols"][-2],
        report_config["right"]["cols"][-2],
    )
    formula_exists = {
        column: worksheet[f"{column}{current_total_row}"].data_type == "f"
        for column in formula_columns
    }

    for merged_range in total_merges:
        worksheet.merged_cells.ranges.remove(merged_range)

    row_delta = target_total_row - current_total_row
    rows_removed = max(0, -row_delta)
    if row_delta < 0:
        worksheet.delete_rows(target_total_row, amount=rows_removed)
    elif row_delta > 0:
        worksheet.insert_rows(current_total_row, amount=row_delta)
        if current_total_row > data_start_row:
            _copy_row_style(
                worksheet,
                source_row=current_total_row - 1,
                start_row=current_total_row,
                end_row=target_total_row - 1,
                end_column=len(
                    report_config["left"]["cols"]
                    + report_config["right"]["cols"]
                ),
            )

    for min_col, min_row, max_col, max_row in total_merge_bounds:
        worksheet.merge_cells(
            start_row=min_row + row_delta,
            start_column=min_col,
            end_row=max_row + row_delta,
            end_column=max_col,
        )

    data_end_row = target_total_row - 1
    for column, existed in formula_exists.items():
        if existed:
            worksheet[f"{column}{target_total_row}"] = (
                f"=SUM({column}{data_start_row}:{column}{data_end_row})"
            )

    for side_name in ("left", "right"):
        side = report_config[side_name]
        for sequence, row in enumerate(
            range(side["data_start_row"], target_total_row),
            start=1,
        ):
            worksheet[f"{side['cols'][0]}{row}"] = sequence
            _clear_cells(worksheet, [row], side["cols"][1:])
    return rows_removed


def _normalize_detail_rows(
    worksheet: Worksheet,
    report_config: dict,
    target_max_row: int,
) -> int:
    """Resize a single detail report and reset its blank numbered rows."""
    data_start_row = report_config["data_start_row"]
    if target_max_row < data_start_row:
        raise NextTemplateError(
            f"{worksheet.title} 目标总行数不得小于数据起始行 {data_start_row}"
        )

    current_max_row = worksheet.max_row
    rows_removed = max(0, current_max_row - target_max_row)
    if current_max_row > target_max_row:
        _discard_merges_in_deleted_rows(worksheet, target_max_row)
        worksheet.delete_rows(
            target_max_row + 1,
            amount=current_max_row - target_max_row,
        )
    elif current_max_row < target_max_row:
        worksheet.insert_rows(
            current_max_row + 1,
            amount=target_max_row - current_max_row,
        )
        if current_max_row >= data_start_row:
            _copy_row_style(
                worksheet,
                source_row=current_max_row,
                start_row=current_max_row + 1,
                end_row=target_max_row,
                end_column=len(report_config["cols"]),
            )

    sequence_column = report_config["cols"][0]
    data_columns = report_config["cols"][1:]
    for sequence, row in enumerate(
        range(data_start_row, target_max_row + 1),
        start=1,
    ):
        worksheet[f"{sequence_column}{row}"] = sequence
        _clear_cells(worksheet, [row], data_columns)
    return rows_removed


def _discard_merges_in_deleted_rows(
    worksheet: Worksheet,
    target_max_row: int,
) -> None:
    """Keep retained merges valid while deleting rows after ``target_max_row``.

    A merge that straddles the deletion boundary cannot remain unchanged after
    the rows below the boundary are removed.  Clip only its bottom edge to the
    last retained row; all other retained merges remain untouched.
    """
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row <= target_max_row < merged_range.max_row:
            worksheet.merged_cells.ranges.remove(merged_range)
            worksheet.merge_cells(
                start_row=merged_range.min_row,
                start_column=merged_range.min_col,
                end_row=target_max_row,
                end_column=merged_range.max_col,
            )
        elif merged_range.min_row > target_max_row:
            worksheet.merged_cells.ranges.remove(merged_range)


def _copy_row_style(
    worksheet: Worksheet,
    source_row: int,
    start_row: int,
    end_row: int,
    end_column: int,
) -> None:
    source_height = worksheet.row_dimensions[source_row].height
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = source_height
        for column in range(1, end_column + 1):
            worksheet.cell(row=row, column=column)._style = copy(
                worksheet.cell(row=source_row, column=column)._style
            )


def _clear_manual_data_fills(
    workbook: Workbook,
    config: dict,
    rules: dict,
) -> int:
    """Clear direct-color fills in data rows while retaining theme fills."""
    cleared = 0
    for report_key, rule in rules["reports"].items():
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        for row in _sanitized_data_rows(worksheet, report_config, rule):
            for cell in worksheet[row]:
                if isinstance(cell, MergedCell):
                    continue
                if cell.fill.fill_type and cell.fill.fgColor.type != "theme":
                    cell.fill = PatternFill()
                    cleared += 1
    return cleared


def _sanitized_data_rows(
    worksheet: Worksheet,
    report_config: dict,
    rule: dict,
) -> set[int]:
    return {
        row
        for start_row, end_row in _sanitized_data_intervals(
            worksheet,
            report_config,
            rule,
        )
        for row in range(start_row, end_row + 1)
    }


def _sanitized_data_intervals(
    worksheet: Worksheet,
    report_config: dict,
    rule: dict,
) -> list[tuple[int, int]]:
    kind = rule.get("kind")
    if kind == "double_detail":
        start_row, total_row = _double_detail_bounds(worksheet, report_config)
        return [(start_row, total_row - 1)]
    if kind == "detail":
        return [(report_config["data_start_row"], worksheet.max_row)]
    if kind == "park":
        return [
            (sub_table["data_start_row"], sub_table["data_end_row"])
            for sub_table in report_config["sub_tables"]
        ]
    return [(report_config["data_start_row"], report_config["data_end_row"])]


def _clear_cells(
    worksheet: Worksheet,
    rows: Iterable[int],
    columns: Iterable[str],
    *,
    preserve_formula_columns: set[str] | None = None,
    preserve_all_formulas: bool = False,
) -> int:
    preserve_formula_columns = preserve_formula_columns or set()
    cleared = 0
    for row in rows:
        for column in columns:
            cell = worksheet[f"{column}{row}"]
            if isinstance(cell, MergedCell):
                continue
            if cell.data_type == "f" and (
                preserve_all_formulas or column in preserve_formula_columns
            ):
                continue
            if cell.value is not None:
                cleared += 1
            cell.value = None
    return cleared


def _expand_column_specs(specs: Iterable[str]) -> set[str]:
    columns: set[str] = set()
    for spec in specs:
        if not isinstance(spec, str) or not spec.strip():
            raise NextTemplateError(f"无效列配置: {spec!r}")
        parts = [part.strip().upper() for part in spec.split(":")]
        if len(parts) == 1:
            column_index_from_string(parts[0])
            columns.add(parts[0])
            continue
        if len(parts) != 2:
            raise NextTemplateError(f"无效列范围: {spec!r}")
        columns.update(_columns_between(parts[0], parts[1]))
    return columns


def _columns_between(start_column: str, end_column: str) -> set[str]:
    start = column_index_from_string(start_column)
    end = column_index_from_string(end_column)
    if start > end:
        raise NextTemplateError(f"列范围起止颠倒: {start_column}:{end_column}")
    return {get_column_letter(index) for index in range(start, end + 1)}


def _find_total_row(
    worksheet: Worksheet,
    data_start_row: int,
    sequence_column: str,
    configured_total_row: int | None,
) -> int:
    for row in range(worksheet.max_row, data_start_row - 1, -1):
        value = worksheet[f"{sequence_column}{row}"].value
        if isinstance(value, str) and value.strip() in _TOTAL_LABELS:
            return row
    if isinstance(configured_total_row, int) and configured_total_row > data_start_row:
        return configured_total_row
    raise NextTemplateError(
        f"{worksheet.title} 未找到 {sequence_column} 列合计行"
    )


def _count_source_period_markers(
    workbook: Workbook,
    current_period: QuarterContext,
    config: dict,
    rules: dict,
) -> int:
    markers = {
        current_period.label,
        current_period.end_date,
        _dotted_date(current_period),
    }
    count = 0
    for report_key in rules["reports"]:
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        for row in _period_text_rows(report_key, report_config, worksheet):
            for cell in worksheet[row]:
                if isinstance(cell.value, str) and any(
                    marker in cell.value for marker in markers
                ):
                    count += 1
    return count


def _replace_period_headers(
    workbook: Workbook,
    current_period: QuarterContext,
    target_period: QuarterContext,
    config: dict,
    rules: dict,
) -> int:
    replacements = _period_replacements(current_period, target_period)
    pattern = re.compile(
        "|".join(
            re.escape(source)
            for source in sorted(replacements, key=len, reverse=True)
        )
    )
    replaced = 0
    for report_key in rules["reports"]:
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        for row in _period_text_rows(report_key, report_config, worksheet):
            for cell in worksheet[row]:
                if not isinstance(cell.value, str) or cell.data_type == "f":
                    continue
                updated, count = pattern.subn(
                    lambda match: replacements[match.group(0)],
                    cell.value,
                )
                if count:
                    cell.value = updated
                    replaced += count
                    logger.debug(
                        "  [期间文字更新] %s!%s: %s → %s",
                        worksheet.title,
                        cell.coordinate,
                        current_period.code,
                        target_period.code,
                    )
    return replaced


def _period_replacements(
    current_period: QuarterContext,
    target_period: QuarterContext,
) -> dict[str, str]:
    pairs = (
        (current_period, target_period),
        (current_period.previous(), target_period.previous()),
        (current_period.previous_year(), target_period.previous_year()),
    )
    replacements: dict[str, str] = {}
    for source, target in pairs:
        replacements[source.label] = target.label
        replacements[source.end_date] = target.end_date
        replacements[_dotted_date(source)] = _dotted_date(target)
        replacements[_month_day(source)] = _month_day(target)

    replacements.update(
        {
            _cumulative_range(current_period, include_end_year=True):
                _cumulative_range(target_period, include_end_year=True),
            _cumulative_range(current_period, include_end_year=False):
                _cumulative_range(target_period, include_end_year=False),
        }
    )
    return {
        source: target
        for source, target in replacements.items()
        if source and source != target
    }


def _period_text_rows(
    report_key: str,
    report_config: dict,
    worksheet: Worksheet,
) -> set[int]:
    if report_key == "report5":
        return set(range(1, report_config["left"]["data_start_row"]))
    if report_key == "report7":
        rows = {1}
        for sub_table in report_config["sub_tables"]:
            header_row = sub_table["header_row"]
            rows.update({header_row - 1, header_row})
        return rows
    if report_key == "report8":
        return set(range(1, report_config["data_start_row"]))
    data_rows = set(
        range(
            report_config["data_start_row"],
            report_config["data_end_row"] + 1,
        )
    )
    return set(range(1, worksheet.max_row + 1)) - data_rows


def _capture_structure(workbook: Workbook) -> dict[str, tuple[str, ...]]:
    return {
        worksheet.title: tuple(
            sorted(str(merged_range) for merged_range in worksheet.merged_cells.ranges)
        )
        for worksheet in workbook.worksheets
    }


def _validate_structure_unchanged(
    workbook: Workbook,
    before: dict[str, tuple[str, ...]],
) -> None:
    after = _capture_structure(workbook)
    if tuple(before) != tuple(after):
        raise NextTemplateError("生成过程中工作表名称或顺序发生变化")
    for sheet_name, merged_ranges in before.items():
        if after[sheet_name] != merged_ranges:
            raise NextTemplateError(f"生成过程中合并区域发生变化: {sheet_name}")


def _validate_sanitized_template(
    workbook: Workbook,
    config: dict,
    rules: dict,
) -> None:
    clear_manual_fills = rules.get("clear_manual_fills", False)
    for report_key, rule in rules["reports"].items():
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        target_max_row = rule.get("target_max_row")
        if target_max_row is not None and worksheet.max_row != target_max_row:
            raise NextTemplateError(
                f"{worksheet.title} 清理后应为 {target_max_row} 行，"
                f"实际为 {worksheet.max_row} 行"
            )
        if target_max_row is not None:
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.max_row > target_max_row:
                    raise NextTemplateError(
                        f"{worksheet.title} 缩行后仍引用已删除行的合并区域: "
                        f"{merged_range}"
                    )

        if clear_manual_fills:
            for row in _sanitized_data_rows(worksheet, report_config, rule):
                for cell in worksheet[row]:
                    if (
                        not isinstance(cell, MergedCell)
                        and cell.fill.fill_type
                        and cell.fill.fgColor.type != "theme"
                    ):
                        raise NextTemplateError(
                            f"{worksheet.title} 数据区仍包含遗留填充: {cell.coordinate}"
                        )


def _validate_no_comments(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    raise NextTemplateError(
                        f"下季度模板仍包含批注: {worksheet.title}!{cell.coordinate}"
                    )


def _enable_excel_recalculation(workbook: Workbook) -> None:
    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def _month_day(period: QuarterContext) -> str:
    return period.end_date.split("年", 1)[1]


def _dotted_date(period: QuarterContext) -> str:
    month_day = _month_day(period).replace("月", ".").replace("日", "")
    return f"{period.year}.{month_day}"


def _cumulative_range(
    period: QuarterContext,
    *,
    include_end_year: bool,
) -> str:
    end = period.end_date if include_end_year else _month_day(period)
    return f"{period.year}年1月1日-{end}"

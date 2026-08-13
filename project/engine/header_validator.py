"""Read-only ownership header structure checks."""

import logging
from dataclasses import dataclass

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class HeaderSignature:
    """Minimal structural signature for one report header block."""

    min_column: int | None
    max_column: int | None
    merged_ranges: tuple[tuple[int, int, int, int], ...]

    @property
    def column_range(self) -> str:
        if self.min_column is None or self.max_column is None:
            return "<空>"
        return (
            f"{get_column_letter(self.min_column)}:"
            f"{get_column_letter(self.max_column)}"
        )


def detect_ownership_header_mismatches(
    template_workbook,
    ownership_data: dict,
    config: dict,
    report_ids: tuple[int, ...] | list[int] | None = None,
) -> dict[tuple[str, int], str]:
    """Detect source sheets whose header structure differs from the template.

    The first version deliberately compares only the effective header column
    range and merged-cell coordinates. It does not interpret header text,
    modify workbooks, or attempt to repair a mismatch.
    """
    selected_report_ids = tuple(report_ids or range(1, 9))
    mismatches: dict[tuple[str, int], str] = {}
    checked_count = 0

    for report_id in selected_report_ids:
        report_config = config["reports"][f"report{report_id}"]
        sheet_name = report_config["sheet_name"]
        template_ws = template_workbook[sheet_name]
        header_start_row, header_end_row = _get_header_rows(report_config)
        template_signature = _build_header_signature(
            template_ws,
            header_start_row,
            header_end_row,
        )

        for owner_key, owner_data in ownership_data.items():
            workbook = owner_data["workbook"]
            if sheet_name not in workbook.sheetnames:
                reason = f"缺少工作表 '{sheet_name}'"
                mismatches[(owner_key, report_id)] = reason
                continue

            checked_count += 1
            source_signature = _build_header_signature(
                workbook[sheet_name],
                header_start_row,
                header_end_row,
            )
            reasons = _compare_signatures(template_signature, source_signature)
            if not reasons:
                continue

            reason = "；".join(reasons)
            mismatches[(owner_key, report_id)] = reason

    logger.debug(
        "权属表格式检查完成：检查 %s 个工作表，需核对 %s 项",
        checked_count,
        len(mismatches),
    )
    return mismatches


def _get_header_rows(report_config: dict) -> tuple[int, int]:
    """Return the configured primary header block for one report."""
    if "sub_tables" in report_config:
        first_sub_table = report_config["sub_tables"][0]
        header_end_row = int(first_sub_table["header_row"])
    elif "data_start_row" in report_config:
        header_end_row = int(report_config["data_start_row"]) - 1
    elif "left" in report_config:
        header_end_row = int(report_config["left"]["data_start_row"]) - 1
    else:
        raise ValueError(
            f"无法确定工作表 '{report_config.get('sheet_name', '<未知>')}' 的表头结束行"
        )

    header_start_row = int(
        report_config.get("header_start_row", header_end_row)
    )
    if header_start_row < 1 or header_start_row > header_end_row:
        raise ValueError(
            f"工作表 '{report_config.get('sheet_name', '<未知>')}' 的表头范围无效: "
            f"{header_start_row}-{header_end_row}"
        )
    return header_start_row, header_end_row


def _build_header_signature(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
) -> HeaderSignature:
    """Build an effective column range and relative merge signature."""
    occupied_columns: set[int] = set()
    merged_ranges = []

    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if _has_content(cell.value):
                occupied_columns.add(cell.column)

    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        occupied_columns.update(
            range(merged_range.min_col, merged_range.max_col + 1)
        )
        merged_ranges.append(
            (
                merged_range.min_row - start_row,
                merged_range.max_row - start_row,
                merged_range.min_col,
                merged_range.max_col,
            )
        )

    return HeaderSignature(
        min(occupied_columns) if occupied_columns else None,
        max(occupied_columns) if occupied_columns else None,
        tuple(sorted(merged_ranges)),
    )


def _compare_signatures(
    template_signature: HeaderSignature,
    source_signature: HeaderSignature,
) -> list[str]:
    """Describe structural differences without attempting a repair."""
    reasons = []
    if (
        template_signature.min_column,
        template_signature.max_column,
    ) != (
        source_signature.min_column,
        source_signature.max_column,
    ):
        reasons.append(
            f"表头有效列范围不一致（模板 {template_signature.column_range}，"
            f"权属表 {source_signature.column_range}）"
        )
    if template_signature.merged_ranges != source_signature.merged_ranges:
        reasons.append(
            "表头合并结构不一致"
            f"（模板 {len(template_signature.merged_ranges)} 处，"
            f"权属表 {len(source_signature.merged_ranges)} 处）"
        )
    return reasons


def _has_content(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True

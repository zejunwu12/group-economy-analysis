"""Matcher for unit/园区 name matching in data files"""

import logging
from typing import Any

from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _normalize(name: str) -> str:
    """归一化名称：移除所有空白字符（空格、换行、制表符等）。"""
    return "".join(name.split())


def find_unit_row(
    ws: Worksheet,
    unit_name: str,
    col_letter: str = "B",
    data_start_row: int = 1,
) -> int | None:
    """在数据文件的指定列中查找单位名称，返回行号。

    匹配规则：归一化（去除空白字符）后精确比对。

    Args:
        ws: 数据文件的工作表
        unit_name: 配置中的单位名称（如 "中侨集团"）
        col_letter: 搜索的列字母，默认 "B"
        data_start_row: 从第几行开始搜索，默认 1

    Returns:
        匹配到的行号，未找到返回 None

    Raises:
        ValueError: 搜索列或起始行不合法
    """
    norm_target = _normalize(unit_name)
    if not norm_target:
        return None

    if data_start_row < 1:
        raise ValueError("data_start_row 必须大于等于 1")

    try:
        column_index = column_index_from_string(col_letter.strip().upper())
    except (AttributeError, ValueError) as exc:
        raise ValueError(f"无效的列字母: {col_letter!r}") from exc

    for row in range(data_start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        cell_value = cell.value

        # 跳过空单元格
        if cell_value is None:
            # 检查是否属于合并单元格区域
            cell_value = _get_merged_value(ws, cell)

        if cell_value is None:
            continue

        if _normalize(str(cell_value)) == norm_target:
            return row

    return None


def _get_merged_value(ws: Worksheet, cell: Any) -> Any | None:
    """如果单元格属于合并区域且值为 None，从合并区域左上角获取值。"""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = ws.cell(
                row=merged_range.min_row, column=merged_range.min_col
            )
            return top_left.value
    return None

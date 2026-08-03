"""Validator: template consistency, total value checking, anomaly detection"""

import logging
import re

from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_SUM_FORMULA_PATTERN = re.compile(
    r"^=SUM\(\s*(?P<start_col>[A-Z]+)(?P<start_row>\d+)\s*:"
    r"\s*(?P<end_col>[A-Z]+)(?P<end_row>\d+)\s*\)$",
    re.IGNORECASE,
)

_FIXED_REPORT_IDS = frozenset({1, 2, 3, 4, 6})
_AREA_HEADER_MARKERS = ("面积", "㎡", "平方米", "m²", "m2")


def validate_totals(
    workbook: Workbook,
    config: dict,
    report_ids: tuple[int, ...] = (1, 2, 3, 4, 6),
    tolerance: float = 0.01,
) -> dict:
    """校验固定报表的 SUM 合计公式是否覆盖完整数据区。

    openpyxl 不会重算 Excel 公式，因此本函数不读取公式缓存结果。
    而是分别计算完整数据区与 SUM 公式引用区的数值合计，二者差异
    超过 tolerance 时说明公式可能漏行、错列或引用范围错误。
    """
    details = []

    for report_id in report_ids:
        report_key = f"report{report_id}"
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        start_row = report_config["data_start_row"]
        end_row = report_config["data_end_row"]
        total_row = report_config["total_row"]
        start_col = column_index_from_string(report_config["data_start_col"])
        end_col = column_index_from_string(report_config["data_end_col"])

        for column in range(start_col, end_col + 1):
            column_letter = get_column_letter(column)
            expected_sum = _sum_numeric_cells(worksheet, column, start_row, end_row)
            total_value = worksheet.cell(row=total_row, column=column).value

            if _is_sum_formula(total_value):
                detail = _validate_sum_formula(
                    worksheet,
                    report_id,
                    column_letter,
                    total_value,
                    expected_sum,
                    tolerance,
                )
            elif _is_formula(total_value):
                detail = {
                    "report_id": report_id,
                    "sheet_name": worksheet.title,
                    "column": column_letter,
                    "status": "skipped",
                    "formula": total_value,
                    "reason": "非 SUM 的计算型公式，不按列合计校验",
                }
            elif expected_sum != 0:
                detail = {
                    "report_id": report_id,
                    "sheet_name": worksheet.title,
                    "column": column_letter,
                    "status": "warning",
                    "formula": None,
                    "expected_sum": expected_sum,
                    "reason": "数据区存在数值，但合计行没有 SUM 公式",
                }
            else:
                detail = {
                    "report_id": report_id,
                    "sheet_name": worksheet.title,
                    "column": column_letter,
                    "status": "skipped",
                    "formula": total_value,
                    "reason": "数据区没有数值，跳过合计校验",
                }

            details.append(detail)

    summary = {
        "passed": sum(detail["status"] == "passed" for detail in details),
        "warnings": sum(detail["status"] == "warning" for detail in details),
        "skipped": sum(detail["status"] == "skipped" for detail in details),
        "details": details,
    }
    logger.info(
        f"合计值校验完成: 通过 {summary['passed']} 项，"
        f"需核对 {summary['warnings']} 项，未校验 {summary['skipped']} 项"
    )
    for detail in details:
        if detail["status"] == "warning":
            logger.warning(_format_warning(detail))
    return summary


def detect_anomalies(
    workbook: Workbook,
    config: dict,
    report_ids: tuple[int, ...] = (1, 2, 3, 4, 5, 6, 7, 8),
    max_area: float = 1_000_000,
    area_columns: dict[int | str, tuple[str, ...] | list[str]] | None = None,
) -> dict:
    """检测写入后数据区中的负面积、超过面积阈值和关键字段为空。

    面积列优先读取 report 配置中的 ``area_columns``，也可以通过
    area_columns 参数显式指定，例如 ``{4: ("C", "D", "E", "F", "G")}``；
    未配置的新报表结构才根据数据区上方表头自动识别。
    本函数只生成报告和日志，不修改工作簿内容。
    """
    if isinstance(max_area, bool) or not isinstance(max_area, (int, float)):
        raise TypeError("max_area 必须是正数")
    if max_area <= 0:
        raise ValueError("max_area 必须大于 0")

    anomalies = []
    scanned_area_cells = 0
    area_column_report = {}

    for report_id in report_ids:
        report_key = f"report{report_id}"
        report_config = config["reports"][report_key]
        worksheet = workbook[report_config["sheet_name"]]
        zones = _get_anomaly_zones(worksheet, report_id, report_config)
        explicit_columns = _get_explicit_area_columns(area_columns, report_id)
        if explicit_columns is None:
            explicit_columns = _get_configured_area_columns(report_config)
        detected_columns = set()

        for zone in zones:
            zone_area_columns = (
                explicit_columns & set(zone["columns"])
                if explicit_columns is not None
                else _detect_area_columns(
                    worksheet,
                    zone["columns"],
                    zone["start_row"],
                )
            )
            detected_columns.update(zone_area_columns)

            for row in range(zone["start_row"], zone["end_row"] + 1):
                empty_anomaly = _detect_empty_row(
                    worksheet,
                    report_id,
                    report_config,
                    zone,
                    row,
                )
                if empty_anomaly is not None:
                    anomalies.append(empty_anomaly)

                for column_letter in sorted(zone_area_columns):
                    cell = worksheet[f"{column_letter}{row}"]
                    value = cell.value
                    if not _is_numeric_value(value):
                        continue

                    scanned_area_cells += 1
                    if value < 0:
                        anomalies.append(
                            _make_cell_anomaly(
                                report_id,
                                worksheet,
                                cell.coordinate,
                                value,
                                "negative_area",
                                "面积字段为负数",
                            )
                        )
                    elif value > max_area:
                        anomalies.append(
                            _make_cell_anomaly(
                                report_id,
                                worksheet,
                                cell.coordinate,
                                value,
                                "oversized_area",
                                f"面积超过阈值 {max_area:g} ㎡",
                            )
                        )

        area_column_report[report_key] = sorted(detected_columns)

    summary = {
        "total": len(anomalies),
        "negative_area": sum(
            item["anomaly_type"] == "negative_area" for item in anomalies
        ),
        "oversized_area": sum(
            item["anomaly_type"] == "oversized_area" for item in anomalies
        ),
        "empty_key_data": sum(
            item["anomaly_type"] == "empty_key_data" for item in anomalies
        ),
        "scanned_area_cells": scanned_area_cells,
        "area_columns": area_column_report,
        "details": anomalies,
    }
    logger.info(
        f"数据完整性和异常值检测完成: 共发现 {summary['total']} 项，"
        f"负面积 {summary['negative_area']} 项，"
        f"超过面积阈值 {summary['oversized_area']} 项，"
        f"关键字段均为空 {summary['empty_key_data']} 项"
    )
    for anomaly in anomalies:
        logger.warning(_format_anomaly(anomaly))
    return summary


def _get_anomaly_zones(
    worksheet: Worksheet,
    report_id: int,
    report_config: dict,
) -> list[dict]:
    """按报表结构返回待检测的数据区及关键字段列。"""
    if report_id in _FIXED_REPORT_IDS:
        columns = _column_letters(
            report_config["data_start_col"], report_config["data_end_col"]
        )
        excluded_columns = {
            column.upper() for column in report_config.get("formula_columns", {})
        } | {
            column.upper() for column in report_config.get("text_columns", [])
        }
        return [{
            "start_row": report_config["data_start_row"],
            "end_row": report_config["data_end_row"],
            "columns": columns,
            "empty_columns": [
                column for column in columns if column not in excluded_columns
            ],
            "identity_column": report_config.get("b_col"),
            "only_mapped_rows": True,
        }]

    if report_id == 5:
        zones = []
        for side in ("left", "right"):
            side_config = report_config[side]
            columns = [column.upper() for column in side_config["cols"]]
            zones.append({
                "start_row": side_config["data_start_row"],
                "end_row": _find_report5_data_end(worksheet, side_config),
                "columns": columns,
                "empty_columns": columns[1:],
                "identity_column": columns[0],
                "only_mapped_rows": False,
            })
        return zones

    if report_id == 7:
        columns = _column_letters(
            report_config["data_start_col"], report_config["data_end_col"]
        )
        return [{
            "start_row": sub_table["data_start_row"],
            "end_row": sub_table["data_end_row"],
            "columns": columns,
            # D:T 是园区表的实际经营数据；A:C 为名称/主体，U:V 是说明和备注。
            "empty_columns": _column_letters("D", "T"),
            "identity_column": "A",
            "only_mapped_rows": True,
        } for sub_table in report_config["sub_tables"]]

    if report_id == 8:
        columns = [column.upper() for column in report_config["cols"]]
        return [{
            "start_row": report_config["data_start_row"],
            "end_row": _find_report8_data_end(worksheet, columns),
            "columns": columns,
            "empty_columns": columns[1:],
            "identity_column": columns[0],
            "only_mapped_rows": False,
        }]

    raise ValueError(f"不支持的报表编号: {report_id}")


def _detect_empty_row(
    worksheet: Worksheet,
    report_id: int,
    report_config: dict,
    zone: dict,
    row: int,
) -> dict | None:
    """识别应有数据但关键输入字段全部为空的行。"""
    if zone["end_row"] < zone["start_row"]:
        return None

    if zone["only_mapped_rows"]:
        if report_id == 7:
            if not _is_report7_mapped_row(report_config, row):
                return None
        elif row not in report_config["row_mapping"]:
            return None

    identity_column = zone["identity_column"]
    identity = _get_cell_or_merged_value(worksheet, f"{identity_column}{row}")
    if not _has_content(identity):
        return None

    if any(
        _has_non_formula_content(worksheet[f"{column}{row}"].value)
        for column in zone["empty_columns"]
    ):
        return None

    return {
        "report_id": report_id,
        "sheet_name": worksheet.title,
        "cell": f"{identity_column}{row}",
        "row": row,
        "column": identity_column,
        "value": identity,
        "anomaly_type": "empty_key_data",
        "message": "关键字段均为空",
    }


def _detect_area_columns(
    worksheet: Worksheet,
    columns: list[str],
    data_start_row: int,
) -> set[str]:
    """从数据区上方的多行表头中识别面积列。"""
    area_columns = set()
    for column in columns:
        header_text = " ".join(
            str(_get_cell_or_merged_value(worksheet, f"{column}{row}"))
            for row in range(1, data_start_row)
            if _has_content(_get_cell_or_merged_value(worksheet, f"{column}{row}"))
        ).lower()
        if any(marker in header_text for marker in _AREA_HEADER_MARKERS):
            area_columns.add(column)
    return area_columns


def _get_explicit_area_columns(
    area_columns: dict[int | str, tuple[str, ...] | list[str]] | None,
    report_id: int,
) -> set[str] | None:
    """读取可选的显式面积列配置。"""
    if area_columns is None:
        return None
    columns = area_columns.get(report_id)
    if columns is None:
        columns = area_columns.get(f"report{report_id}")
    if columns is None:
        return set()
    return {str(column).upper() for column in columns}


def _get_configured_area_columns(report_config: dict) -> set[str] | None:
    """读取单张报表可选的 area_columns 配置。"""
    columns = report_config.get("area_columns")
    if columns is None:
        return None
    return {str(column).upper() for column in columns}


def _find_report5_data_end(worksheet: Worksheet, side_config: dict) -> int:
    """查找双栏报表5移动后的合计行，以适配动态扩展和收缩。"""
    sequence_column = side_config["cols"][0].upper()
    for row in range(side_config["data_start_row"], worksheet.max_row + 1):
        value = worksheet[f"{sequence_column}{row}"].value
        if _normalize_text(value) == "合计":
            return row - 1
    return min(side_config["data_end_row"], worksheet.max_row)


def _find_report8_data_end(worksheet: Worksheet, columns: list[str]) -> int:
    """查找报表8动态明细区的最后一个实际记录行。"""
    for row in range(worksheet.max_row, 0, -1):
        if any(_has_content(worksheet[f"{column}{row}"].value) for column in columns):
            return row
    return 0


def _is_report7_mapped_row(report_config: dict, row: int) -> bool:
    return any(row in sub_table["row_mapping"] for sub_table in report_config["sub_tables"])


def _column_letters(start_column: str, end_column: str) -> list[str]:
    return [
        get_column_letter(column)
        for column in range(
            column_index_from_string(start_column),
            column_index_from_string(end_column) + 1,
        )
    ]


def _get_cell_or_merged_value(worksheet: Worksheet, cell_ref: str) -> object | None:
    """读取普通单元格或其所在合并区域左上角的值。"""
    cell = worksheet[cell_ref]
    if cell.value is not None:
        return cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _has_content(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _has_non_formula_content(value: object) -> bool:
    return _has_content(value) and not _is_formula(value)


def _is_numeric_value(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return "".join(str(value).split())


def _make_cell_anomaly(
    report_id: int,
    worksheet: Worksheet,
    cell_ref: str,
    value: int | float,
    anomaly_type: str,
    message: str,
) -> dict:
    column_letter = re.match(r"[A-Z]+", cell_ref).group(0)
    row = int(re.search(r"\d+", cell_ref).group(0))
    return {
        "report_id": report_id,
        "sheet_name": worksheet.title,
        "cell": cell_ref,
        "row": row,
        "column": column_letter,
        "value": value,
        "anomaly_type": anomaly_type,
        "message": message,
    }


def _format_anomaly(anomaly: dict) -> str:
    sheet_name = _display_sheet_name(
        anomaly["report_id"], anomaly["sheet_name"]
    )
    return (
        f"报表{anomaly['report_id']} {sheet_name}｜"
        f"位置：{anomaly['cell']}｜{anomaly['message']}｜"
        f"标识值：{anomaly['value']}"
    )


def _validate_sum_formula(
    worksheet,
    report_id: int,
    column_letter: str,
    formula: str,
    expected_sum: float,
    tolerance: float,
) -> dict:
    """校验单个 SUM 公式的引用范围与完整数据区合计是否一致。"""
    match = _SUM_FORMULA_PATTERN.match(formula.replace("$", ""))
    if match is None:
        return {
            "report_id": report_id,
            "sheet_name": worksheet.title,
            "column": column_letter,
            "status": "skipped",
            "formula": formula,
            "reason": "SUM 公式格式不支持自动解析",
        }

    start_col = match.group("start_col").upper()
    end_col = match.group("end_col").upper()
    start_row = int(match.group("start_row"))
    end_row = int(match.group("end_row"))
    if start_col != column_letter or end_col != column_letter:
        return {
            "report_id": report_id,
            "sheet_name": worksheet.title,
            "column": column_letter,
            "status": "warning",
            "formula": formula,
            "expected_sum": expected_sum,
            "reason": "SUM 公式引用了其他列或多列范围",
        }

    formula_sum = _sum_numeric_cells(
        worksheet,
        column_index_from_string(column_letter),
        start_row,
        end_row,
    )
    difference = formula_sum - expected_sum
    detail = {
        "report_id": report_id,
        "sheet_name": worksheet.title,
        "column": column_letter,
        "status": "passed" if abs(difference) <= tolerance else "warning",
        "formula": formula,
        "expected_sum": expected_sum,
        "formula_sum": formula_sum,
        "difference": difference,
        "formula_range": f"{column_letter}{start_row}:{column_letter}{end_row}",
    }
    if detail["status"] == "warning":
        detail["reason"] = "SUM 公式范围计算结果与完整数据区合计不一致"
    return detail


def _sum_numeric_cells(worksheet, column: int, start_row: int, end_row: int) -> float:
    """累加指定列和行范围的数值，忽略文本、公式与 bool。"""
    total = 0.0
    for row in range(start_row, end_row + 1):
        value = worksheet.cell(row=row, column=column).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            total += value
    return total


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_sum_formula(value: object) -> bool:
    return _is_formula(value) and value.upper().startswith("=SUM(")


def _format_warning(detail: dict) -> str:
    """生成便于日志阅读的合计校验告警。"""
    sheet_name = _display_sheet_name(detail["report_id"], detail["sheet_name"])
    location = f"报表{detail['report_id']} {sheet_name}｜{detail['column']}列"
    return f"{location}｜{detail['reason']}"


def _display_sheet_name(report_id: int, sheet_name: str) -> str:
    """去掉工作表名称中重复的报表编号，仅用于日志显示。"""
    prefix = f"报表{report_id} "
    if sheet_name.startswith(prefix):
        return sheet_name[len(prefix):]
    return sheet_name

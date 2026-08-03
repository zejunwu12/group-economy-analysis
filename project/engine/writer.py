"""Data writer with merged cell protection"""

import logging
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter

from engine.comments import CommentCopyStats, copy_source_comment
from engine.matcher import find_unit_row

logger = logging.getLogger(__name__)

_TOTAL_ROW_LABELS = frozenset({"合计", "汇总", "总计"})


class MergedCellProtectionError(Exception):
    """写入前后合并单元格结构发生变化。"""


def write_report_fixed(
    template_ws: Worksheet,
    report_config: dict,
    ownership_data: dict,
    config: dict,
    report_id: int,
    comment_stats: CommentCopyStats | None = None,
) -> int:
    """将固定行数报表的权属数据写入汇总表。

    流程：
    1. 遍历 row_mapping，取出行号和单位名称
    2. 通过 unit_to_owner 找到该单位所属的权属文件
    3. 在权属文件中用 B 列查找该单位的数据行
    4. 将该行数据写入汇总表对应行

    数值列采用“非数值归零”策略：仅数值原样写入，空值、``/`` 等非数值
    均写入 0；``text_columns`` 中的文本列仍保留来源原值。``formula_columns``
    只用于识别公式列，公式内容始终保留模板原值，不由本函数重写。

    Args:
        template_ws: 汇总表模板的工作表
        report_config: 该报表的配置（含 row_mapping, data_start_col, data_end_col 等）
        ownership_data: load_ownership_files() 的返回值
        config: 配置字典（含 unit_to_owner）
        report_id: 报表编号（用于日志）

    Returns:
        成功写入的行数（不含集团本部等无数据源的行）

    Raises:
        MergedCellProtectionError: 写入前后合并区域坐标发生变化
    """
    merged_ranges_before = get_merged_ranges(template_ws)
    row_mapping = report_config["row_mapping"]
    data_start_col = report_config["data_start_col"]
    data_end_col = report_config["data_end_col"]
    sheet_name = report_config["sheet_name"]
    formula_columns = {
        str(column).upper()
        for column in report_config.get("formula_columns", {})
    }
    formula_number_format = report_config.get("formula_number_format")
    text_columns = {
        str(column).upper() for column in report_config.get("text_columns", [])
    }
    difference_check_columns = {
        str(column).upper()
        for column in report_config.get("difference_check_columns", [])
    }

    ownership_files = config["ownership_files"]

    start_col_idx = column_index_from_string(data_start_col)
    end_col_idx = column_index_from_string(data_end_col)
    written_count = 0

    _warn_unconfigured_source_units(
        template_ws,
        report_config,
        ownership_data,
        config,
        report_id,
    )

    if formula_number_format:
        format_rows = list(row_mapping)
        total_row = report_config.get("total_row")
        if isinstance(total_row, int):
            format_rows.append(total_row)
        for row_num in format_rows:
            for column in formula_columns:
                format_cell = _get_writable_cell(
                    template_ws,
                    row_num,
                    column_index_from_string(column),
                )
                if format_cell is not None:
                    format_cell.number_format = formula_number_format

    for row_num, unit_name in row_mapping.items():
        # 查找该单位所属的权属
        owner_key = _resolve_owner_key(
            template_ws,
            row_num,
            unit_name,
            report_config,
            config,
        )
        if owner_key is None:
            logger.warning(
                f"  报表{report_id} 第{row_num}行: 单位 '{unit_name}' "
                f"无法确定所属权属，跳过"
            )
            continue

        # 检查是否有数据文件
        owner_config = ownership_files[owner_key]
        if owner_config["file"] is None:
            logger.debug(f"  报表{report_id} 第{row_num}行: {unit_name} 无数据文件，跳过")
            continue

        # 获取该权属的 workbook
        owner_data = ownership_data.get(owner_key)
        if owner_data is None:
            logger.warning(
                f"  报表{report_id} 第{row_num}行: 权属 '{owner_key}' 未加载，跳过"
            )
            continue

        # 在权属文件中找到该单位的数据行
        try:
            src_ws = owner_data["workbook"][sheet_name]
        except KeyError:
            logger.warning(
                f"  报表{report_id} 第{row_num}行: "
                f"权属 '{owner_key}' 缺少工作表 '{sheet_name}'，跳过"
            )
            continue

        src_row = find_unit_row(
            src_ws,
            unit_name,
            data_start_row=report_config.get("data_start_row", 1),
        )
        if src_row is None:
            logger.warning(
                f"  报表{report_id} 第{row_num}行: "
                f"在 {owner_key} 的 {sheet_name} 中未找到 '{unit_name}'，跳过"
            )
            continue

        # 文本列保留原值；数值列只保留数值，其他来源值统一归零。
        # 即使来源值为空，也要显式写入，以免模板或重复运行时留下旧值。
        copied_cells = 0
        numeric_cells = 0
        zero_filled_cells = 0
        for col_idx in range(start_col_idx, end_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            if column_letter in formula_columns:
                continue

            source_cell = src_ws.cell(row=src_row, column=col_idx)
            src_value = source_cell.value
            target_cell = _get_writable_cell(template_ws, row_num, col_idx)
            if target_cell is None:
                continue

            if target_cell.data_type == "f":
                continue

            if (
                column_letter in difference_check_columns
                and target_cell.value != src_value
            ):
                _warn_source_difference(
                    template_ws=template_ws,
                    source_ws=src_ws,
                    target_cell=target_cell,
                    source_cell=source_cell,
                    report_config=report_config,
                    report_id=report_id,
                    owner_key=owner_key,
                    owner_data=owner_data,
                    unit_name=unit_name,
                )

            if column_letter in text_columns:
                target_cell.value = src_value
            elif _is_numeric_value(src_value):
                target_cell.value = src_value
                numeric_cells += 1
            else:
                target_cell.value = 0
                zero_filled_cells += 1
            copy_source_comment(
                source_cell,
                target_cell,
                report_id=report_id,
                owner_key=owner_key,
                source_ws=src_ws,
                target_ws=template_ws,
                stats=comment_stats,
            )
            copied_cells += 1

        logger.debug(
            f"  报表{report_id} 第{row_num}行: {unit_name} ← "
            f"{owner_key} 第{src_row}行，写入 {copied_cells} 个直接填入单元格，"
            f"数值 {numeric_cells} 个，非数值归零 {zero_filled_cells} 个"
        )
        written_count += 1

    validate_merged_ranges(template_ws, merged_ranges_before)
    return written_count


def get_merged_ranges(ws: Worksheet) -> frozenset[str]:
    """返回工作表当前全部合并区域的坐标快照。"""
    return frozenset(str(merged_range) for merged_range in ws.merged_cells.ranges)


def validate_merged_ranges(
    ws: Worksheet,
    expected_ranges: frozenset[str],
) -> None:
    """确认合并区域坐标与写入前完全一致。"""
    actual_ranges = get_merged_ranges(ws)
    if actual_ranges == expected_ranges:
        logger.debug(
            f"  {ws.title}: 合并单元格结构保持不变 "
            f"({len(actual_ranges)}处)"
        )
        return

    removed = sorted(expected_ranges - actual_ranges)
    added = sorted(actual_ranges - expected_ranges)
    raise MergedCellProtectionError(
        f"{ws.title} 合并单元格结构发生变化；"
        f"缺失: {removed or '无'}；新增: {added or '无'}"
    )


def _get_writable_cell(
    ws: Worksheet,
    row: int,
    column: int,
) -> Cell | None:
    """返回可写逻辑单元格；合并区域非左上角返回 None。"""
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    return cell


def _is_numeric_value(value: object) -> bool:
    """判断来源值是否为可直接写入数值列的数值。"""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _warn_source_difference(
    template_ws: Worksheet,
    source_ws: Worksheet,
    target_cell: Cell,
    source_cell: Cell,
    report_config: dict,
    report_id: int,
    owner_key: str,
    owner_data: dict,
    unit_name: str,
) -> None:
    """在覆盖模板前记录配置列中的来源数据差异。"""
    column_letter = target_cell.column_letter
    header = _get_column_header(
        template_ws,
        column_letter,
        report_config.get("data_start_row", target_cell.row),
    )
    filename = (
        owner_data.get("filename")
        or report_config.get("source_filename")
        or f"{owner_key}.xlsx"
    )
    sheet_name = _display_sheet_name(report_id, template_ws.title)
    logger.warning(
        f"  [环比/同比数据差异检测] 报表{report_id} {sheet_name}｜"
        f"权属：{owner_key}｜单位：{unit_name}｜"
        f"字段：{header}（{column_letter}列）｜"
        f"模板原值：{template_ws.title}!{target_cell.coordinate}="
        f"{_format_difference_value(target_cell.value)}｜"
        f"权属文件值：{filename}/{source_ws.title}!{source_cell.coordinate}="
        f"{_format_difference_value(source_cell.value)}"
    )


def _get_column_header(
    worksheet: Worksheet,
    column_letter: str,
    data_start_row: int,
) -> str:
    """读取数据区上一行的列标题，兼容纵向或横向合并单元格。"""
    header_row = max(1, data_start_row - 1)
    value = _get_cell_or_merged_value(
        worksheet,
        f"{column_letter}{header_row}",
    )
    if value is None:
        return f"{column_letter}列"
    return " ".join(str(value).split())


def _format_difference_value(value: object) -> str:
    """将差异值格式化为单行、可辨识的日志文本。"""
    if value is None:
        return "<空值>"
    if isinstance(value, str):
        return repr(" ".join(value.split()))
    return repr(value)


def _resolve_owner_key(
    template_ws: Worksheet,
    row_num: int,
    unit_name: str,
    report_config: dict,
    config: dict,
) -> str | None:
    """优先按单位名称反查权属，失败时使用模板 A 列集团名称。"""
    owner_key = config["unit_to_owner"].get(unit_name)
    if owner_key is not None:
        return owner_key

    owner_mapping = report_config.get("owner_mapping", {})
    owner_key = owner_mapping.get(unit_name)
    if owner_key in config["ownership_files"]:
        return owner_key

    owner_col = report_config.get("a_col")
    if not owner_col:
        return None

    owner_value = _get_cell_or_merged_value(template_ws, f"{owner_col}{row_num}")
    if owner_value is None:
        return None

    normalized_owner = _normalize_name(str(owner_value))
    for candidate in config["ownership_files"]:
        if _normalize_name(candidate) == normalized_owner:
            return candidate
    return None


def _get_cell_or_merged_value(ws: Worksheet, cell_ref: str) -> object | None:
    """读取单元格值；若为空且位于合并区域，读取区域左上角。"""
    cell = ws[cell_ref]
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
    return None


def _normalize_name(name: str) -> str:
    """移除名称中的空白字符。"""
    return "".join(name.split())


def _warn_unconfigured_source_units(
    template_ws: Worksheet,
    report_config: dict,
    ownership_data: dict,
    config: dict,
    report_id: int,
) -> None:
    """提示源表中未出现在本报表配置里的单位，但不参与写入。"""
    sheet_name = report_config["sheet_name"]
    expected_by_owner = _get_expected_names_by_owner(
        template_ws,
        report_config,
        config,
    )
    owner_names_by_owner = _get_owner_names_by_owner(
        template_ws,
        report_config,
        config,
    )

    for owner_key, owner_data in ownership_data.items():
        try:
            source_ws = owner_data["workbook"][sheet_name]
        except KeyError:
            continue

        expected_names = expected_by_owner.get(owner_key, set())
        extras = _find_unconfigured_source_units(
            source_ws,
            report_config,
            expected_names,
            owner_names_by_owner.get(
                owner_key,
                {_normalize_name(owner_key)},
            ),
        )
        if not extras:
            continue

        details = "、".join(
            f"'{unit_name}'（源表第{row_num}行）"
            for row_num, unit_name in extras
        )
        logger.warning(
            f"  [未配置单位检测] 报表{report_id}｜权属：{owner_key}｜"
            f"发现 {len(extras)} 个未配置单位：{details}；"
            "处理结果：未写入汇总表，请核对配置和模板"
        )


def _display_sheet_name(report_id: int, sheet_name: str) -> str:
    """去掉工作表名称中重复的报表编号，仅用于日志显示。"""
    prefix = f"报表{report_id} "
    if sheet_name.startswith(prefix):
        return sheet_name[len(prefix):]
    return sheet_name


def _get_expected_names_by_owner(
    template_ws: Worksheet,
    report_config: dict,
    config: dict,
) -> dict[str, set[str]]:
    """按权属整理本报表 row_mapping 中允许写入的单位名称。"""
    expected_by_owner: dict[str, set[str]] = {}
    for row_num, unit_name in report_config["row_mapping"].items():
        owner_key = _resolve_owner_key(
            template_ws,
            row_num,
            unit_name,
            report_config,
            config,
        )
        if owner_key is None:
            continue
        expected_by_owner.setdefault(owner_key, set()).add(
            _normalize_name(unit_name)
        )
    return expected_by_owner


def _get_owner_names_by_owner(
    template_ws: Worksheet,
    report_config: dict,
    config: dict,
) -> dict[str, set[str]]:
    """整理权属键及其在模板 A 列中使用的名称。"""
    owner_names_by_owner = {
        owner_key: {_normalize_name(owner_key)}
        for owner_key in config["ownership_files"]
    }
    owner_col = report_config.get("a_col", "A")

    for row_num, unit_name in report_config["row_mapping"].items():
        owner_key = _resolve_owner_key(
            template_ws,
            row_num,
            unit_name,
            report_config,
            config,
        )
        if owner_key is None:
            continue
        owner_value = _get_cell_or_merged_value(
            template_ws,
            f"{owner_col}{row_num}",
        )
        if owner_value is not None:
            owner_names_by_owner.setdefault(owner_key, set()).add(
                _normalize_name(str(owner_value))
            )

    return owner_names_by_owner


def _find_unconfigured_source_units(
    source_ws: Worksheet,
    report_config: dict,
    expected_names: set[str],
    expected_owner_names: set[str],
) -> list[tuple[int, str]]:
    """扫描源表数据区，返回未配置单位的 ``(行号, 名称)``。"""
    start_row = report_config.get("data_start_row", 1)
    owner_col = report_config.get("a_col", "A")
    unit_col = report_config.get("b_col", "B")
    extras: list[tuple[int, str]] = []

    for row_num in range(start_row, source_ws.max_row + 1):
        owner_value = _get_cell_or_merged_value(
            source_ws,
            f"{owner_col}{row_num}",
        )
        unit_value = _get_cell_or_merged_value(
            source_ws,
            f"{unit_col}{row_num}",
        )

        normalized_owner = (
            _normalize_name(str(owner_value)) if owner_value is not None else ""
        )
        normalized_unit = (
            _normalize_name(str(unit_value)) if unit_value is not None else ""
        )
        if _is_total_row_label(normalized_owner) or _is_total_row_label(
            normalized_unit
        ):
            break
        if (
            not normalized_unit
            or normalized_unit in expected_names
            or normalized_owner not in expected_owner_names
        ):
            continue

        extras.append((row_num, str(unit_value).strip()))

    return extras


def _is_total_row_label(value: str) -> bool:
    """识别“合计”“资产情况总计”等数据区结束标记。"""
    return value in _TOTAL_ROW_LABELS or any(
        value.endswith(label) for label in _TOTAL_ROW_LABELS
    )


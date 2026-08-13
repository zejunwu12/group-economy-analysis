"""Report 5 handler: double-column detail appending"""

import logging
from copy import copy

from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from engine.comments import CommentCopyStats, copy_source_comment
from engine.writer import get_merged_ranges, validate_merged_ranges

logger = logging.getLogger(__name__)


def process_report5(
    template_ws: Worksheet,
    ownership_data: dict,
    report_config: dict,
    comment_stats: CommentCopyStats | None = None,
    *,
    excluded_owners: set[str] | None = None,
) -> dict[str, int]:
    """汇总报表5左右两栏明细，必要时扩展模板数据行。"""
    sheet_name = report_config["sheet_name"]
    left_config = report_config["left"]
    right_config = report_config["right"]

    left_records = _extract_records(
        ownership_data,
        sheet_name,
        left_config,
        "左侧拟改造",
        excluded_owners=excluded_owners,
    )
    right_records = _extract_records(
        ownership_data,
        sheet_name,
        right_config,
        "右侧正在改造",
        excluded_owners=excluded_owners,
    )

    required_rows = max(len(left_records), len(right_records))
    data_end_row, total_row, expanded_rows, removed_rows = _resize_data_area(
        template_ws,
        report_config,
        required_rows,
    )

    _clear_data_area(
        template_ws,
        left_config["data_start_row"],
        data_end_row,
        left_config["cols"] + right_config["cols"],
    )
    _write_records(
        template_ws, left_records, left_config["cols"], left_config["data_start_row"],
        report_id=5, comment_stats=comment_stats,
    )
    _write_records(
        template_ws, right_records, right_config["cols"], right_config["data_start_row"],
        report_id=5, comment_stats=comment_stats,
    )

    logger.info(
        f"报表5汇总完成: 左侧 {len(left_records)} 条，"
        f"右侧 {len(right_records)} 条，合计行第 {total_row} 行"
    )
    return {
        "left_count": len(left_records),
        "right_count": len(right_records),
        "expanded_rows": expanded_rows,
        "removed_rows": removed_rows,
        "data_end_row": data_end_row,
        "total_row": total_row,
    }


def _extract_records(
    ownership_data: dict,
    sheet_name: str,
    side_config: dict,
    side_label: str,
    *,
    excluded_owners: set[str] | None = None,
) -> list[dict]:
    """提取一侧有效明细；序号列不参与有效性判断。"""
    columns = [column_index_from_string(column) for column in side_config["cols"]]
    records = []
    excluded_owners = excluded_owners or set()

    for owner_key, owner_data in ownership_data.items():
        if owner_key in excluded_owners:
            continue
        workbook = owner_data["workbook"]
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"报表5: 权属 '{owner_key}' 缺少工作表 '{sheet_name}'，跳过")
            continue

        worksheet = workbook[sheet_name]
        owner_count = 0
        for row in range(side_config["data_start_row"], worksheet.max_row + 1):
            sequence_value = worksheet.cell(row=row, column=columns[0]).value
            if _normalize(sequence_value) == "合计":
                break

            values = [worksheet.cell(row=row, column=column).value for column in columns]
            if not any(_has_content(value) for value in values[1:]):
                continue

            records.append(
                {
                    "owner": owner_key,
                    "worksheet": worksheet,
                    "source_row": row,
                    "values": values,
                }
            )
            owner_count += 1

        logger.debug(f"  报表5 {owner_key} {side_label}: {owner_count} 条")

    return records


def _resize_data_area(
    worksheet: Worksheet,
    report_config: dict,
    required_rows: int,
) -> tuple[int, int, int, int]:
    """按左右两栏最大明细数扩展或收缩数据区。"""
    left_config = report_config["left"]
    right_config = report_config["right"]
    data_start_row = left_config["data_start_row"]
    data_end_row = left_config["data_end_row"]
    total_row = left_config["total_row"]

    if (
        right_config["data_start_row"] != data_start_row
        or right_config["data_end_row"] != data_end_row
        or right_config["total_row"] != total_row
    ):
        raise ValueError("报表5左右两栏的数据区和合计行必须一致")

    capacity = data_end_row - data_start_row + 1
    row_delta = required_rows - capacity
    expanded_rows = max(0, row_delta)
    removed_rows = max(0, -row_delta)
    if row_delta == 0:
        return data_end_row, total_row, 0, 0

    merged_before = get_merged_ranges(worksheet)
    total_merges = [
        merged_range
        for merged_range in worksheet.merged_cells.ranges
        if merged_range.min_row <= total_row <= merged_range.max_row
    ]
    total_merge_bounds = [
        (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        for merged_range in total_merges
    ]
    total_merge_names = {str(merged_range) for merged_range in total_merges}

    formula_columns = [
        left_config["cols"][-2],
        right_config["cols"][-2],
    ]
    formula_exists = {
        column: _is_formula(worksheet[f"{column}{total_row}"].value)
        for column in formula_columns
    }

    for merged_range in total_merges:
        worksheet.unmerge_cells(str(merged_range))

    if row_delta > 0:
        worksheet.insert_rows(total_row, amount=expanded_rows)
        new_total_row = total_row + expanded_rows
        new_data_end_row = new_total_row - 1
        _copy_data_row_style(
            worksheet,
            source_row=data_end_row,
            start_row=total_row,
            end_row=new_data_end_row,
            start_col=1,
            end_col=10,
        )
    else:
        delete_start_row = data_start_row + required_rows
        worksheet.delete_rows(delete_start_row, amount=removed_rows)
        new_total_row = total_row - removed_rows
        new_data_end_row = new_total_row - 1

    shifted_total_merges = set()
    for min_col, min_row, max_col, max_row in total_merge_bounds:
        shifted_range = (
            f"{get_column_letter(min_col)}{min_row + row_delta}:"
            f"{get_column_letter(max_col)}{max_row + row_delta}"
        )
        worksheet.merge_cells(shifted_range)
        shifted_total_merges.add(shifted_range)

    for column, existed in formula_exists.items():
        if existed:
            if required_rows == 0:
                worksheet[f"{column}{new_total_row}"] = "=0"
            else:
                worksheet[f"{column}{new_total_row}"] = (
                    f"=SUM({column}{data_start_row}:{column}{new_data_end_row})"
                )

    expected_merges = frozenset(
        (merged_before - total_merge_names) | shifted_total_merges
    )
    validate_merged_ranges(worksheet, expected_merges)
    if row_delta > 0:
        logger.warning(
            f"报表5明细超出模板容量 {expanded_rows} 行，"
            f"数据区已扩展至第 {new_data_end_row} 行"
        )
    else:
        logger.info(
            f"报表5明细少于模板容量 {removed_rows} 行，"
            f"数据区已收缩至第 {new_data_end_row} 行"
        )
    return new_data_end_row, new_total_row, expanded_rows, removed_rows


def _copy_data_row_style(
    worksheet: Worksheet,
    source_row: int,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """将最后一个预置数据行的样式复制给新增数据行。"""
    source_height = worksheet.row_dimensions[source_row].height
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = source_height
        for column in range(start_col, end_col + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=row, column=column)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)


def _clear_data_area(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    columns: list[str],
) -> None:
    """清空左右两栏数据值，保留单元格样式。"""
    column_indices = [column_index_from_string(column) for column in columns]
    for row in range(start_row, end_row + 1):
        for column in column_indices:
            worksheet.cell(row=row, column=column).value = None


def _write_records(
    worksheet: Worksheet,
    records: list[dict],
    columns: list[str],
    start_row: int,
    *,
    report_id: int,
    comment_stats: CommentCopyStats | None = None,
) -> None:
    """写入一侧明细，并从1开始重新连续编号。"""
    column_indices = [column_index_from_string(column) for column in columns]
    for sequence, record in enumerate(records, start=1):
        target_row = start_row + sequence - 1
        worksheet.cell(row=target_row, column=column_indices[0]).value = sequence
        source_ws = record["worksheet"]
        source_row = record["source_row"]
        for column in column_indices[1:]:
            source_cell = source_ws.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=target_row, column=column)
            target_cell.value = source_cell.value
            copy_source_comment(
                source_cell,
                target_cell,
                report_id=report_id,
                owner_key=record["owner"],
                source_ws=source_ws,
                target_ws=worksheet,
                stats=comment_stats,
            )


def _has_content(value: object) -> bool:
    """判断明细字段是否包含有效内容；数字0属于有效值。"""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _normalize(value: object) -> str:
    """移除文本中的空白字符。"""
    if value is None:
        return ""
    return "".join(str(value).split())


def _is_formula(value: object) -> bool:
    """判断单元格值是否为公式。"""
    return isinstance(value, str) and value.startswith("=")
